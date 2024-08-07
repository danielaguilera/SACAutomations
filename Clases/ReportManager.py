from tkinter import *
from tkinter import ttk

from PyPDF2 import PdfMerger
from Clases.MailSender import MailSender
from Clases.PDFGenerator import PDFGenerator
from Clases.SACConnector import SACConnector
from Clases.SACSenderJob import SACSenderJob
from Utils.GlobalFunctions import *
from Utils.Metadata import *
from Clases.Servicio import Servicio
from tkinter import messagebox
import os
import shutil
import traceback
import io
import threading
# from PIL import ImageTk, Image
import fitz

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Border, Side, Alignment, NamedStyle

class ReportManager:
    def __init__(self, container):
        self.container = container
        self.mode = SEND
        self.toplevel = Toplevel(background='seashell4')
        self.toplevel.title(string='Reportes a enviar')
        self.toplevel.protocol("WM_DELETE_WINDOW", self.onClosingWindow)
        self.toplevel.wm_state('zoomed')
        self.thumbnailFrame = Frame(master=self.toplevel)
        self.thumbnailFrame.pack(side=RIGHT)
        self.reporteImage = None
        self.thumbnail = Label(master=self.thumbnailFrame)
        self.sacConnector: SACConnector = SACConnector()
        self.reportFrame = Frame(master=self.toplevel)
        self.reportFrame.pack(expand=True, fill=BOTH)
        self.reportColumns = ['Destinatario', 'Email', '# Boleta', 'ID Mapsa', 'Beneficiario', 'Cliente', 'Deudor', 'Monto Total (CLP)']
        self.columnWidths = [200, 150, 50, 150, 100, 100, 100]
        self.reportTable = ttk.Treeview(master=self.reportFrame, columns=self.reportColumns, show='headings', height=10)
        for columnName in self.reportColumns:
            self.reportTable.heading(columnName, text=columnName)
            self.reportTable.column(columnName, width=150)
        self.reportTable.pack(expand=True, fill=BOTH, anchor=CENTER, padx=5)
        self.reportTable.bind('<<TreeviewSelect>>', self.displayThumbnail)
        
        self.actionFrame = Frame(master=self.toplevel)
        self.actionFrame.pack(expand=True, fill=BOTH)

        self.sendButton = Button(master=self.actionFrame, text='Enviar todo', width=40, height=1, font=('Helvetica bold', 15), fg = 'black', bg='RoyalBlue1', command=self.triggerSendAllReports)
        self.sendButton.pack(expand=False, fill=BOTH)   

        self.sendDestinatarioButton = Button(master=self.actionFrame, text='Enviar todo del destinatario', width=40, height=1, font=('Helvetica bold', 15), fg = 'black', bg='lawngreen', command=self.triggerSendDestinatarioReports)
        self.sendDestinatarioButton.pack(expand=False, fill=BOTH)      

        self.deleteButton = Button(master=self.actionFrame, text='Eliminar reporte', width=40, height=1, font=('Helvetica bold', 15), fg = 'black', bg='indian red', command=self.triggerDeleteReport)
        self.deleteButton.pack(expand=False, fill=BOTH)
        
        self.toplevel.grab_set()
        self.container.master.withdraw()
        
        self.getReports()
        
    def onClosingWindow(self):
        self.container.master.deiconify()
        self.container.master.wm_state('zoomed')
        self.toplevel.destroy()
    
    def getReports(self):
        if not os.path.exists(DELIVEREDDATAPATH):
            return
        self.reportTable.delete(*self.reportTable.get_children())
        nombreDestinatario: str
        for nombreDestinatario in os.listdir(DELIVEREDDATAPATH):
            self.reportTable.insert('', END, iid=nombreDestinatario, values=(nombreDestinatario,'','','','','','',''), open=True)
            reportsFound: list = [dirName for dirName in os.listdir(f'{DELIVEREDDATAPATH}/{nombreDestinatario}') if dirName[0] != 'R']
            for dirName in reportsFound:
                if dirName != 'Documento.pdf':
                    numBoleta, idMapsa = [int(x) for x in dirName.strip().split('_')]
                    with open(f'{DELIVEREDDATAPATH}/{nombreDestinatario}/{numBoleta}_{idMapsa}/Data_{numBoleta}.txt') as file:
                        data = file.readline().strip().split(';')
                        nombreDestinatario = data[0]
                        mailDestinatario = data[1]
                        numBoleta = data[2]
                        idMapsa = data[3]
                        nombreBeneficiario = data[4]
                        nombreCliente = data[5]
                        nombreDeudor = data[6]
                        montoTotal = data[7]
                        tableData = ('-','-',numBoleta,idMapsa,nombreBeneficiario,nombreCliente,nombreDeudor,montoTotal)
                        self.reportTable.insert(nombreDestinatario, END, values=tableData, open=True)
                    self.reportTable.item(nombreDestinatario, values=(nombreDestinatario, mailDestinatario,'','','','','',''))
    
    def getRendicionNumberFromMatrix(self, nombreDestinatario: str, nombreCliente: str):
        excelMatrixRoot = f'{DELIVEREDDATAPATH}/{nombreDestinatario}/Rendición {nombreCliente}.xlsx'
        if not os.path.exists(excelMatrixRoot):
            return 0
        workbook = openpyxl.load_workbook(excelMatrixRoot)
        sheet = workbook.active
        cell = sheet.cell(row = 1, column = 8)
        label: str = cell.value
        return int(label.split(' ')[2].split('/')[0])
    
    def displayThumbnail(self, key=None):
        data = self.reportTable.item(self.reportTable.selection()[0])['values']
        if not(data[2] and data[3]):
            return
        numBoleta: int = int(data[2])
        idMapsa: int = int(data[3])
        nombreDestinatario: str = self.reportTable.item(self.reportTable.parent(self.reportTable.selection()[0]))['values'][0]
        
        zoom_x = 1.0  # horizontal zoom
        zoom_y = 1.0  # vertical zoom
        mat = fitz.Matrix(zoom_x, zoom_y)  # zoom factor 2 in each dimension

        filename = f'{DELIVEREDDATAPATH}/{nombreDestinatario}/{numBoleta}_{idMapsa}/Reporte_{numBoleta}.pdf' 
        
        doc = fitz.open(filename)  # open document
        for page in doc:  # iterate through the pages
            pix = page.get_pixmap(matrix=mat)  # render page to an image
            pix.save("thumbnail.png")  # store image as a PNG
        doc.close()
            
        self.reporteImage = PhotoImage(file='thumbnail.png')
        self.thumbnail.config(image=self.reporteImage)
        self.thumbnail.pack()
        
    def triggerDeleteReport(self):
        if not self.reportTable.focus():
            return
        data = self.reportTable.item(self.reportTable.selection()[0])['values']
        if not(data[2] and data[3]):
            return
        if not checkInternetConnection():
            messagebox.showerror(title='Error', message='No hay conexión a Internet.')
            return
        
        if getCacheFiles():
            task: RunningTask = getCacheFiles()[0]
            if task.script == SENDING_ACTIVITY:
                messagebox.showerror(title='Error', message=f'{task.user} está enviando reportes.\n Por favor intente en unos segundos más.')
                return            
            elif task.script == DELETING_ACTIVITY:
                messagebox.showerror(title='Error', message=f'{task.user} está eliminando reportes.\n Por favor intente en unos segundos más.')      
                return
            elif task.script == ADDING_ACTIVITY:
                messagebox.showerror(title='Error', message=f'{task.user} está añadiendo reportes.\n Por favor intente en unos segundos más.')
                return
        
        if not messagebox.askyesno(title='Aviso', message='¿Estás segur@ de que quieres borrar los datos para esta boleta?'):
            return
        p = threading.Thread(target=self.deleteReport)
        p.start()
    
    def deleteReport(self):        
        if not self.reportTable.focus():
            return
        data = self.reportTable.item(self.reportTable.selection()[0])['values']
        if not(data[2] and data[3]):
            return
        
        # Marcar actividad:
        createCacheFile(user=self.container.user, script=DELETING_ACTIVITY)
        
        # Mostrar la ventana emergente de carga
        loadingWindow = Toplevel(self.container.master)
        loadingWindow.grab_set()
        loadingWindow.title("Cargando...")
            
        # Añadir animación de carga
        loadingLabel = ttk.Label(loadingWindow, text="Borrando... Por favor no cerrar la ventana")
        loadingLabel.pack(padx=20, pady=20)
        progressBar = ttk.Progressbar(loadingWindow, mode="determinate")
        progressBar.pack(padx=20, pady=10)
        progressBar.start()
        
        numBoleta: int = int(data[2])
        idMapsa: int = int(data[3])
        nombreDestinatario: str = self.reportTable.item(self.reportTable.parent(self.reportTable.selection()[0]))['values'][0]
        nombreCliente: str = data[5]
        numeroRendicion: int = self.sacConnector.getRendicionNumber(nombreDestinatario=nombreDestinatario, nombreCliente=nombreCliente)
        self.sacConnector.deleteBoletaData(numBoleta=numBoleta, idMapsa=idMapsa)
        deleteIfExists(f'{DELIVEREDDATAPATH}/{nombreDestinatario}/{numBoleta}_{idMapsa}')
        rows = self.sacConnector.getClienteMatrixRows(nombreDestinatario=nombreDestinatario, nombreCliente=nombreCliente)
        if rows:
            fileGenerator: PDFGenerator = PDFGenerator()
            fileGenerator.updateExcelMatrix(nombreDestinatario=nombreDestinatario, nombreCliente=nombreCliente)
        else:
            deleteFileIfExists(f'{DELIVEREDDATAPATH}/{nombreDestinatario}/Rendición_{nombreCliente}_{numeroRendicion}.xlsx')
        deleteIfEmpty(f'{DELIVEREDDATAPATH}/{nombreDestinatario}')
        deleteIfEmpty(f'{DELIVEREDDATAPATH}')
        
        loadingWindow.destroy()
        
        messagebox.showinfo(title='INFO', message='Reporte borrado')
        with open(ACTIVITYLOGFILE, 'a') as file:
            with open('Params/mail_data.txt') as credentials:
                username, password = credentials.readline().strip().split(',')
                server = SMTPSERVERGYD
                port = SMTPPORTGYD
                mailSender: MailSender = MailSender(senderUsername=username, senderPassword=password, smtpServer=server, smtpPort=port)
                mailSender.sendMessage(receiverAddress='draguilera@uc.cl', mailSubject=f'{"[DEMO] " if self.mode == "demo" else ""}Actividad - {datetime.now()}', mailContent=f'{self.container.user} eliminó la boleta #{numBoleta}, caso {idMapsa}')
            file.write(f'{str(datetime.now())}: {self.container.user} eliminó los archivos de boleta a enviar (NUMERO BOLETA: {numBoleta} - ID MAPSA: {idMapsa}) del destinatario {nombreDestinatario}\n')
            
        self.toplevel.destroy()
        self.container.master.deiconify()
        self.container.master.wm_state('zoomed')
        
        removeCacheFile(user=self.container.user, script=DELETING_ACTIVITY)
        
    def resetForm(self):
        self.toplevel.destroy()
        self.toplevel.update()
        reportManager: ReportManager = ReportManager(container=self.container)
        
    def triggerSendDestinatarioReports(self):
        if not self.reportTable.focus():
            return
        data = self.reportTable.item(self.reportTable.selection()[0])['values']
        if data[2]:
            return
        if not checkInternetConnection():
            messagebox.showerror(title='Error', message='No hay conexión a Internet.')
            return
        nombreDestinatario: str = data[0]

        if getCacheFiles():
            task: RunningTask = getCacheFiles()[0]
            print(task.script)
            if task.script == SENDING_ACTIVITY:
                messagebox.showerror(title='Error', message=f'{task.user} está enviando reportes.\n Por favor intente en unos segundos más.')
                return            
            elif task.script == DELETING_ACTIVITY:
                messagebox.showerror(title='Error', message=f'{task.user} está eliminando reportes.\n Por favor intente en unos segundos más.')      
                return
            elif task.script == ADDING_ACTIVITY:
                messagebox.showerror(title='Error', message=f'{task.user} está añadiendo reportes.\n Por favor intente en unos segundos más.')
                return
        
        if not messagebox.askyesno(title='Aviso', message=f'Se enviarán todas las boletas de esta semana para {nombreDestinatario}.\nEsto puede tardar unos minutos.\n¿Deseas continuar?'):
            return
        p = threading.Thread(target=self.sendDestinatarioReports)
        p.start()
        
    def triggerSendAllReports(self):
        if not checkInternetConnection():
            messagebox.showerror(title='Error', message='No hay conexión a Internet.')
            return
        
        if getCacheFiles():
            task: RunningTask = getCacheFiles()[0]
            print(task.script)
            if task.script == SENDING_ACTIVITY:
                messagebox.showerror(title='Error', message=f'{task.user} está enviando reportes.\n Por favor intente en unos segundos más.')
                return            
            elif task.script == DELETING_ACTIVITY:
                messagebox.showerror(title='Error', message=f'{task.user} está eliminando reportes.\n Por favor intente en unos segundos más.')      
                return
            elif task.script == ADDING_ACTIVITY:
                messagebox.showerror(title='Error', message=f'{task.user} está añadiendo reportes.\n Por favor intente en unos segundos más.')
                return
            elif task.script == VISUALIZING_ACTIVITY and task.user != self.container.user:
                messagebox.showerror(title='Error', message=f'{task.user} está usando la app.\n Por favor intente más tarde.')
                return
        
        if not messagebox.askyesno(title='Aviso', message='Se enviarán todas las boletas guardadas hasta hoy.\nEsto puede tardar unos minutos.\n¿Deseas continuar?'):
            return
        p = threading.Thread(target=self.sendAllReports)
        p.start()

    def sendDestinatarioReports(self):
        if not self.reportTable.focus():
            return
        data = self.reportTable.item(self.reportTable.selection()[0])['values']
        if data[2]:
            return
        nombreDestinatario: str = data[0]
        try:
            # Marcar actividad:
            createCacheFile(user=self.container.user, script=SENDING_ACTIVITY)
            
            # Mostrar la ventana emergente de carga
            loadingWindow = Toplevel(self.container.master)
            loadingWindow.grab_set()
            loadingWindow.title("Cargando...")
            
            # Añadir animación de carga
            loadingLabel = ttk.Label(loadingWindow, text="Enviando... Por favor no cerrar la ventana")
            loadingLabel.pack(padx=20, pady=20)
            progressBar = ttk.Progressbar(loadingWindow, mode="determinate")
            progressBar.pack(padx=20, pady=10)
            progressBar.start()
            
            senderJob: SACSenderJob = SACSenderJob()
            senderJob.generateSingleUnifiedDocument(nombreDestinatario=nombreDestinatario)
            senderJob.sendSingleDestinatarioReports(user=self.container.user, nombreDestinatario=nombreDestinatario)
            loadingWindow.destroy()
            messagebox.showinfo(title='Éxito', message='Reportes enviados')
            self.toplevel.destroy()
            self.container.master.deiconify()
            self.container.master.wm_state('zoomed')
            self.container.clearForm()

        except Exception as e:
            stringBuffer = io.StringIO()
            traceback.print_exc(file=stringBuffer)
            tracebackString = stringBuffer.getvalue()
            stringBuffer.close()
            messagebox.showerror(title='Error', message=tracebackString)
            
        finally:
            removeCacheFile(user=self.container.user, script=SENDING_ACTIVITY)

    def sendAllReports(self):
        try:
            if not os.path.exists(DELIVEREDDATAPATH):
                messagebox.showerror(title='Error', message='No hay reportes para enviar')
                return
            
            # Marcar actividad:
            createCacheFile(user=self.container.user, script=SENDING_ACTIVITY)

            # Mostrar la ventana emergente de carga
            loadingWindow = Toplevel(self.container.master)
            loadingWindow.grab_set()
            loadingWindow.title("Cargando...")

            # Añadir animación de carga
            loadingLabel = ttk.Label(loadingWindow, text="Enviando... Por favor no cerrar la ventana")
            loadingLabel.pack(padx=20, pady=20)
            progressBar = ttk.Progressbar(loadingWindow, mode="determinate")
            progressBar.pack(padx=20, pady=10)
            progressBar.start()

            senderJob: SACSenderJob = SACSenderJob()
            senderJob.generateUnifiedDocument()
            senderJob.sendReports()
            
            loadingWindow.destroy()
            self.container.master.deiconify()
            self.container.master.wm_state('zoomed')
            self.container.clearForm()
            self.toplevel.destroy()
            
            messagebox.showinfo(title='Éxito', message='Reportes enviados')
            
        except Exception as e:
            stringBuffer = io.StringIO()
            traceback.print_exc(file=stringBuffer)
            tracebackString = stringBuffer.getvalue()
            stringBuffer.close()
            messagebox.showerror(title='Error', message=tracebackString)
            
        finally:
            removeCacheFile(user=self.container.user, script=SENDING_ACTIVITY)
        
    def updateUnifiedDocument(self, nombreDestinatario: str):
        if not os.path.exists(DELIVEREDDATAPATH):
            return
        pdfMerger: PdfMerger = PdfMerger()
        for path in os.listdir(path=f'{DELIVEREDDATAPATH}/{nombreDestinatario}'):
            if path != 'Documento.pdf':
                numBoleta, idMapsa = (int(x) for x in path.strip().split('_'))
                reportePath: str = f'{DELIVEREDDATAPATH}/{nombreDestinatario}/{path}/Reporte_{numBoleta}.pdf'
                boletaPath: str = f'{DELIVEREDDATAPATH}/{nombreDestinatario}/{path}/Boleta_{numBoleta}.pdf'
                anexoPath: str = f'{DELIVEREDDATAPATH}/{nombreDestinatario}/{path}/Anexo_{numBoleta}.pdf'
                pdfMerger.append(reportePath)
                pdfMerger.append(boletaPath)
                if os.path.exists(anexoPath):
                    pdfMerger.append(anexoPath)
        pdfMerger.write(f'{DELIVEREDDATAPATH}/{nombreDestinatario}/Documento.pdf')
        if not os.path.exists(f'{GENERATEDREPORTSPATH}/{datetime.today().strftime("%Y-%m-%d")}/{nombreDestinatario}'):
            os.makedirs(f'{GENERATEDREPORTSPATH}/{datetime.today().strftime("%Y-%m-%d")}/{nombreDestinatario}')
        shutil.copy(f'{DELIVEREDDATAPATH}/{nombreDestinatario}/Documento.pdf', f'{GENERATEDREPORTSPATH}/{datetime.today().strftime("%Y-%m-%d")}/{nombreDestinatario}/Documento.pdf')
        pdfMerger.close()

    def generateUnifiedDocument(self):
        if not os.path.exists(DELIVEREDDATAPATH):
            return
        for nombreDestinatario in os.listdir(path=f'{DELIVEREDDATAPATH}'):
            pdfMerger: PdfMerger = PdfMerger()
            for path in os.listdir(path=f'{DELIVEREDDATAPATH}/{nombreDestinatario}'):
                if path != 'Documento.pdf':
                    numBoleta, idMapsa = (int(x) for x in path.strip().split('_'))
                    reportePath: str = f'{DELIVEREDDATAPATH}/{nombreDestinatario}/{path}/Reporte_{numBoleta}.pdf'
                    boletaPath: str = f'{DELIVEREDDATAPATH}/{nombreDestinatario}/{path}/Boleta_{numBoleta}.pdf'
                    anexoPath: str = f'{DELIVEREDDATAPATH}/{nombreDestinatario}/{path}/Anexo_{numBoleta}.pdf'
                    pdfMerger.append(reportePath)
                    pdfMerger.append(boletaPath)
                    if os.path.exists(anexoPath):
                        pdfMerger.append(anexoPath)
            pdfMerger.write(f'{DELIVEREDDATAPATH}/{nombreDestinatario}/Documento.pdf')
            if not os.path.exists(f'{GENERATEDREPORTSPATH}/{datetime.today().strftime("%Y-%m-%d")}/{nombreDestinatario}'):
                os.makedirs(f'{GENERATEDREPORTSPATH}/{datetime.today().strftime("%Y-%m-%d")}/{nombreDestinatario}')
            shutil.copy(f'{DELIVEREDDATAPATH}/{nombreDestinatario}/Documento.pdf', f'{GENERATEDREPORTSPATH}/{datetime.today().strftime("%Y-%m-%d")}/{nombreDestinatario}/Documento.pdf')
        pdfMerger.close()
                
                    