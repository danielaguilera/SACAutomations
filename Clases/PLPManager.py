from Clases.Gestion import Gestion
from Clases.SACConnector import SACConnector
from Clases.MailSender import MailSender
import imaplib
import email
import email.header
from email.message import Message
import sys
import re
from datetime import date, datetime, timedelta, time
from unidecode import unidecode
from Utils.Metadata import *
from Utils.GlobalFunctions import *
from Clases.Caso import Caso
from Clases.Deudor import Deudor
import pytz
import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill

class GYDEmail:
    def __init__(self, msgId: int, sender: str, subject: str, date: str):
        self.msgId: int = msgId
        self.sender: str = sender
        self.subject: str = subject
        self.date: str = date

class PLPRequest:
    def __init__(self, payload: GYDEmail, rutDeudor: str = '', caso: Caso = None):
        self.rutDeudor: str = rutDeudor
        self.caso: Caso = caso
        self.payload: GYDEmail = payload

    def __str__(self):
        return f'Solicitud PLP: {self.rutDeudor}'
    
    def __repr__(self):
        return self.__str__()

class PLPBreached:
    def __init__(self, payload: GYDEmail, deudores: list[Deudor] = []):
        self.deudores: list[Deudor] = deudores
        self.payload: GYDEmail = payload

    def __str__(self):
        return f'PLP Incumplido: {",".join([deudor.rawName for deudor in self.deudores])}'

    def __repr__(self):
        return self.__str__()
    
class JudicialCollection:
    def __init__(self, requestType: str, payload: GYDEmail, rutDeudor: str = '', caso: Caso = None):
        self.rutDeudor: str = rutDeudor
        self.caso: Caso = caso
        self.payload: GYDEmail = payload
        self.requestType: str = requestType
    
    def __str__(self):
        return f'{self.requestType}: {self.rutDeudor}'

class SACRequests:
    def __init__(self, plpRequests: list[PLPRequest], plpBreachedRequests: list[PLPBreached], judicialCollectionRequests: list[JudicialCollection]):
        self.plpRequests = plpRequests
        self.plpBreachedRequests = plpBreachedRequests
        self.judicialCollectionRequests = judicialCollectionRequests

    def __str__(self):
        return '\n'.join([str(elem) for elem in self.plpRequests + self.plpBreachedRequests])
    
class UnMappedRequest:
    def __init__(self, emisor: str, asunto: str, tipoSolicitud: str, nombreDeudor: str, rutDeudor: str):
        self.emisor = emisor
        self.asunto = asunto
        self.tipoSolicitud = tipoSolicitud
        self.nombreDeudor = nombreDeudor
        self.rutDeudor = rutDeudor

class PLPManager:
    def __init__(self):
        self.sacConnector: SACConnector = SACConnector()
        self.smtpServer: str = SMTPSERVERGYD
        self.smtpPort: int = SMTPPORTGYD
        with open(MAILDATA, 'r') as file:
            senderUsername, senderPassword = file.readline().strip().split(',')
        self.mailSender: MailSender = MailSender(senderUsername=senderUsername, 
                                                 senderPassword=senderPassword, 
                                                 smtpServer=self.smtpServer, 
                                                 smtpPort=self.smtpPort)
        self.username: str = senderUsername
        self.password: str = senderPassword
        self.localTimezone = pytz.timezone('America/Santiago')

    def fetchMailData(self, date: datetime = None):
        date = date if date else datetime.now()
        sacRequests: SACRequests = self.getEmails(sinceDatetime=date)
        self.processRequests(sacRequests=sacRequests)

    def sendSummary(self):
        text: str = self.generateSummary()
        self.mailSender.sendPLPSummary(text=text, attachFile=not self.isRequestFileEmpty())
        deleteFileIfExists(PLPREQUESTSPATH)

    def fetchDailyMails(self, date: datetime):
        sacRequests: SACRequests = self.getEmails(sinceDatetime=date)
        self.processRequests(sacRequests=sacRequests)
        text: str = self.generateSummary(date=date)
        self.mailSender.sendPLPSummary(text=text, attachFile=not self.isRequestFileEmpty(), date=date)
        deleteFileIfExists(PLPREQUESTSPATH)

    def getMissingMailIds(self, originalIds: list) -> list:
        if not os.path.exists(PLPREQUESTSPATH):
            return originalIds
        workbook = openpyxl.load_workbook(PLPREQUESTSPATH)
        worksheet = workbook.active
        idValues = [cell.value for cell in worksheet['A'][1:]]
        missingIds = [idNum for idNum in originalIds if int(idNum) not in idValues]
        return missingIds
    
    def touchSolicitudesFile(self):
        if os.path.exists(PLPREQUESTSPATH):
            return
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        headers = ['ID', 'Timestamp', 'Emisor', 'Asunto', 'Tipo', 'Deudor', 'RUT Deudor', 'ID Mapsa', 'Estado Anterior', 'Estado Actual', 'Última modificación']
        for colNum, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=colNum)
            cell.value = header
        workbook.save(PLPREQUESTSPATH)
        
    def decodeHeader(self, text: str) -> str:
        try:
            decodedParts = email.header.decode_header(text)
            decodedSubject = ''
            for part, encoding in decodedParts:
                if isinstance(part, bytes):
                    decodedSubject += part.decode(encoding or 'utf-8')
                else:
                    decodedSubject += part
            return decodedSubject
        except Exception:
            return ''
    
    def convertToLocalTime(self, timeString: str) -> datetime:
        localTimezone = pytz.timezone('America/Santiago')
        formatString = "%a, %d %b %Y %H:%M:%S %z (%Z)"
        dt = datetime.strptime(timeString, formatString)
        dt = dt.astimezone(localTimezone)
        return dt
    
    def isPLPRequest(self, subject: str) -> bool:
        return any(keyWord in subject for keyWord in SOLICITUD_PLP_KEYWORDS) and PLP in subject

    def isPLPBreached(self, subject: str) -> bool:
        return any(keyWord in subject for keyWord in PLP_INCUMPLIDO_KEYWORDS)
    
    def isJudicialCollection(self, subject: str) -> bool:
        return any(keyWord in subject for keyWord in JUDICIAL_COLLECTION_KEYWORDS)
    
    def getRUTFromJudicialCollection(self, text: str) -> str:
        print(text)
        lines: list[str] = text.upper().split('\n')
        for line in lines:
            if 'RUT' in line:
                return correctRUTFormat(line)
        
    def getJudicialCollectionType(self, text: str) -> str:
        if SOLICITUD_CONTINUAR in text:
            return SOLICITUD_CONTINUAR
        elif SOLICITUD_RETIRAR in text:
            return SOLICITUD_RETIRAR
        elif SOLICITUD_SUSPENDER in text:
            return SOLICITUD_SUSPENDER
    
    def isDateFormat(self, word: str) -> bool:
        try:
            datetime.strptime(word, '%d-%m-%Y')
            return True
        except Exception:
            return False
        
    def getDeudoresFromPLPBreachedContent(self, content: str) -> list[Deudor]:
        deudores: list[Deudor] = []

        start = re.search(r'PLP_SECUENCIA', content)
        if not start:
            return []
        end = re.findall(r'PLP[1-9]', content)
        if not end:
            return []
        lastEnd = content.rfind(end[-1])
        croppedContent: str = content[start.start() + 14:lastEnd + 4]

        tokens: list[str] = list(filter(lambda elem: elem, croppedContent.split('\n')))
        rawNames: list[str] = []
        for index, element in enumerate(tokens):
            if self.isDateFormat(element):
                rawNames.append(tokens[index - 2])
        for rawName in rawNames:
            tokenizedName: list[str] = list(filter(lambda word: word and not word.isdigit(), rawName.split(' ')))
            apellidoDeudor: str = ''
            nombreDeudor: str = ''
            if len(tokenizedName) == 2:
                apellidoDeudor = tokenizedName[1]
                nombreDeudor = tokenizedName[0]
            elif len(tokenizedName) >= 3:
                apellidoDeudor = f'{tokenizedName[-2]} {tokenizedName[-1]}'
                nombreDeudor = ' '.join(tokenizedName[0:len(tokenizedName)-2])
            deudor: Deudor = Deudor(rawName=rawName, apellidoDeudor=apellidoDeudor, nombreDeudor=nombreDeudor)
            deudores.append(deudor)
        return deudores

    def getEmails(self, sinceDatetime: datetime = datetime.now(), save: bool = True) -> SACRequests:
        imap = imaplib.IMAP4_SSL(self.smtpServer)
        imap.login(self.username, self.password)
        imap.select("Inbox")
        beforeDatetime: datetime = datetime.now()
        sinceDate: str = sinceDatetime.strftime('%d-%b-%Y')
        beforeDate: str = beforeDatetime.strftime('%d-%b-%Y')
        _, msgnums = imap.search(None, f'(SINCE {sinceDate} BEFORE {beforeDate})')

        plpRequests: list[PLPRequest] = []
        plpBreachedRequests: list[PLPBreached] = []
        judicialCollectionRequests: list[JudicialCollection] = []

        if save:
            msgIds = self.getMissingMailIds(msgnums[0].split())
        else:
            msgIds = msgnums[0].split()

        for msgnum in msgIds:
            _, data = imap.fetch(msgnum, '(RFC822)')
            message: Message = email.message_from_bytes(data[0][1])
            messageSender: str = self.decodeHeader(message.get('From'))
            messageDate: str = message.get('Date')
            messageSubject: str = self.decodeHeader(message.get('Subject'))
            
            messageString: str = ''
            for part in message.walk():
                if part.get_content_type() == "text/plain":
                    messageString += part.as_string()
                        
            if "NO TOMAR EN CUENTA LA REACTIVACIÓN" in messageString.upper() or "NO TOMAR EN CUENTA LA REACTIVACION" in messageString.upper():
                continue
            
            if self.isJudicialCollection(messageSubject.upper()):
                if not save:
                    continue
                requestType: str = self.getJudicialCollectionType(messageSubject.upper())
                gydEmail: GYDEmail = GYDEmail(sender=messageSender,
                                              subject=messageSubject,
                                              date=messageDate,
                                              msgId=int(msgnum))
                
                messageString: str = ''
                for part in message.walk():
                    if part.get_content_type() == "text/plain":
                        messageString += part.as_string()
                try:
                    rutDeudor: str = self.getRUTFromJudicialCollection(messageString.upper())
                    if not rutDeudor:
                        rutDeudor: str = self.getRUTFromJudicialCollection(str(message).upper())
                except Exception:    
                    rutDeudor: str = 'RUT no encontrado'
                    
                casos: list[Caso] = self.sacConnector.getPossibleMapsaCasos(rutDeudor=rutDeudor, active=False)
                caso: Caso = None
                if len(casos) == 1:
                    caso: Caso = casos[0]
                judicialCollectionRequest: JudicialCollection = JudicialCollection(payload=gydEmail,
                                                                                   rutDeudor=rutDeudor,
                                                                                   caso=caso,
                                                                                   requestType=requestType)
                judicialCollectionRequests.append(judicialCollectionRequest)

            elif self.isPLPRequest(messageSubject.upper()):
                if not save:
                    continue
                gydEmail: GYDEmail = GYDEmail(sender=messageSender,
                                              subject=messageSubject,
                                              date=messageDate,
                                              msgId=int(msgnum))
                try:
                    rutDeudor: str = correctRUTFormat(messageSubject)
                except Exception:
                    rutDeudor: str = 'RUT no encontrado'
                casos: list[Caso] = self.sacConnector.getPossibleMapsaCasos(rutDeudor=rutDeudor, active=False)
                caso: Caso = None
                if len(casos) == 1:
                    caso: Caso = casos[0]
                plpRequest: PLPRequest = PLPRequest(payload=gydEmail,
                                                    rutDeudor=rutDeudor,
                                                    caso=caso)
                plpRequests.append(plpRequest)

            elif self.isPLPBreached(messageSubject.upper()):
                if not save:
                    continue
                gydEmail: GYDEmail = GYDEmail(sender=messageSender,
                                              subject=messageSubject,
                                              date=messageDate,
                                              msgId=int(msgnum))

                deudoresFound: list[Deudor] = self.getDeudoresFromPLPBreachedContent(content=messageString)
                plpBreachedRequest: PLPBreached = PLPBreached(payload=gydEmail, deudores=deudoresFound)
                plpBreachedRequests.append(plpBreachedRequest)

        sacRequests: SACRequests = SACRequests(plpRequests=plpRequests, 
                                               plpBreachedRequests=plpBreachedRequests,
                                               judicialCollectionRequests=judicialCollectionRequests)
        imap.close()
        return sacRequests
    
    def processRequests(self, sacRequests: SACRequests):
        self.touchSolicitudesFile()
        processedPLPRequests: list[PLPRequest] = self.processPLPRequests(sacRequests.plpRequests)
        processedPLPBreachedRequests: list[PLPBreached] = self.processPLPBreachedRequests(sacRequests.plpBreachedRequests)
        processedJudicialCollectionRequests: list[JudicialCollection] = self.processJudicialCollectionRequests(sacRequests.judicialCollectionRequests)
        data = []
        for plpRequest in processedPLPRequests:
            if plpRequest.caso:
                data.append([plpRequest.payload.msgId,
                             plpRequest.payload.date, 
                             plpRequest.payload.sender, 
                             plpRequest.payload.subject, 
                             'Solicitud PLP', 
                             plpRequest.caso.nombreDeudor + ' ' + plpRequest.caso.apellidoDeudor, 
                             plpRequest.rutDeudor, 
                             plpRequest.caso.idMapsa, 
                             plpRequest.caso.nombreEstadoAnterior, 
                             plpRequest.caso.nombreEstado,
                             plpRequest.caso.lastGestionDate.date().strftime('%d-%m-%Y') if plpRequest.caso.lastGestionDate else 'N/A'
                            ])
            else:
                data.append([plpRequest.payload.msgId,
                             plpRequest.payload.date, 
                             plpRequest.payload.sender, 
                             plpRequest.payload.subject, 
                             'Solicitud PLP', 
                             'No encontrado', 
                             plpRequest.rutDeudor, 
                             'No encontrado', 
                             'No encontrado', 
                             'No encontrado',
                             'N/A'])
                
        for request in processedJudicialCollectionRequests:
            if request.caso:
                data.append([request.payload.msgId,
                             request.payload.date,
                             request.payload.sender,
                             request.payload.subject,
                             request.requestType,
                             request.caso.nombreDeudor + ' ' + request.caso.apellidoDeudor,
                             request.caso.rutDeudor,
                             request.caso.idMapsa,
                             request.caso.nombreEstadoAnterior,
                             request.caso.nombreEstado,
                             request.caso.lastGestionDate.date().strftime('%d-%m-%Y') if request.caso.lastGestionDate else 'N/A'
                            ])
            else:
                data.append([request.payload.msgId,
                             request.payload.date,
                             request.payload.sender,
                             request.payload.subject,
                             request.requestType,
                             'No encontrado',
                             request.rutDeudor,
                             'No encontrado',
                             'No encontrado',
                             'No encontrado',
                             'N/A'])
            
        for plpBreachedRequest in processedPLPBreachedRequests:
            deudor: Deudor
            for deudor in plpBreachedRequest.deudores:
                if deudor.casoAsociado:
                    data.append([plpBreachedRequest.payload.msgId,
                                 plpBreachedRequest.payload.date,
                                 plpBreachedRequest.payload.sender,
                                 plpBreachedRequest.payload.subject,
                                 'PLP Incumplido',
                                 deudor.rawName,
                                 deudor.casoAsociado.rutDeudor,
                                 deudor.casoAsociado.idMapsa,
                                 deudor.casoAsociado.nombreEstadoAnterior,
                                 deudor.casoAsociado.nombreEstado,
                                 deudor.casoAsociado.lastGestionDate.date().strftime('%d-%m-%Y') if deudor.casoAsociado.lastGestionDate else 'N/A',
                                 ])
                else:
                    data.append([plpBreachedRequest.payload.msgId,
                                 plpBreachedRequest.payload.date,
                                 plpBreachedRequest.payload.sender,
                                 plpBreachedRequest.payload.subject,
                                 'PLP Incumplido',
                                 deudor.rawName,
                                 'No encontrado',
                                 'No encontrado',
                                 'No encontrado',
                                 'No encontrado',
                                 'N/A'
                                 ])
        workbook = openpyxl.load_workbook(PLPREQUESTSPATH)
        worksheet = workbook.active
        alertColor = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
        for rowData in data:
            worksheet.append(rowData)
            newRowIndex = worksheet.max_row
            if rowData[-1] != 'N/A':
                lastDate = datetime.strptime(rowData[-1], '%d-%m-%Y')
                if (datetime.today().date() - lastDate.date()).days < MIN_REQUEST_REPETITION_DELAY:
                    for columnIndex in range(1, len(rowData) + 1):
                        cell = worksheet.cell(row=newRowIndex, column=columnIndex)
                        cell.fill = alertColor
            
        workbook.save(PLPREQUESTSPATH)
            
    def processPLPRequests(self, plpRequests: list[PLPRequest]) -> list[PLPRequest]:
        plpRequest: PLPRequest
        processedPLPRequests: list[PLPRequest] = []
        for plpRequest in plpRequests:
            rutDeudor: str = plpRequest.rutDeudor
            casos: list[Caso] = self.sacConnector.getPossibleMapsaCasos(rutDeudor=rutDeudor, active=False)
            if len(casos) == 1:
                caso: Caso = casos[0]
                caso.lastGestionDate = self.sacConnector.getLastGestionControl(idMapsa=caso.idMapsa)
                plpRequest.caso = caso
                self.sacConnector.setMapsaCasoState(idMapsa=caso.idMapsa, newState=SUSPENDIDO.lower().capitalize())
                gestion: Gestion = Gestion(idJuicio=caso.idMapsa, timestamp=datetime.now(), tipo=PLP, rutDeudor=caso.rutDeudor, nombreDeudor=caso.nombreDeudor + ' ' + caso.apellidoDeudor)
                self.sacConnector.addGestion(gestion=gestion)
                plpRequest.caso.nombreEstado = SUSPENDIDO.lower().capitalize()
            processedPLPRequests.append(plpRequest)
        return processedPLPRequests
    
    def processJudicialCollectionRequests(self, judicialCollectionRequests: list[JudicialCollection]) -> list[JudicialCollection]:
        request: JudicialCollection
        requests: list[JudicialCollection] = judicialCollectionRequests
        processedRequests: list[JudicialCollection] = []
        for request in requests:
            rutDeudor: str = request.rutDeudor
            casos: list[Caso] = self.sacConnector.getPossibleMapsaCasos(rutDeudor=rutDeudor, active=False)
            if len(casos) == 1:
                caso: Caso = casos[0]
                caso.lastGestionDate = self.sacConnector.getLastGestionControl(idMapsa=caso.idMapsa)
                request.caso = caso
                newState: str = JUDICIAL_COLLECTION_STATES[request.requestType]
                self.sacConnector.setMapsaCasoState(idMapsa=caso.idMapsa, newState=newState.lower().capitalize())
                gestion: Gestion = Gestion(idJuicio=caso.idMapsa, timestamp=datetime.now(), tipo=request.requestType, rutDeudor=caso.rutDeudor, nombreDeudor=caso.nombreDeudor + ' ' + caso.apellidoDeudor)
                self.sacConnector.addGestion(gestion=gestion)
                request.caso.nombreEstado = newState.lower().capitalize()
            processedRequests.append(request)
        return processedRequests
                
    def processPLPBreachedRequests(self, plpBreachedRequests: list[PLPBreached]) -> list[PLPBreached]:
        plpBreachedRequest: PLPBreached
        processedPLPBreachedRequests: list[PLPBreached] = []
        for plpBreachedRequest in plpBreachedRequests:
            deudor: Deudor
            mappedDeudores: list[Deudor] = []
            for deudor in plpBreachedRequest.deudores:
                casos: list[Caso] = self.sacConnector.getPossibleMapsaCasos(apellidoDeudor=deudor.apellidoDeudor, nombreDeudor=deudor.nombreDeudor, active=False)
                if len(casos) == 1:
                    caso: Caso = casos[0]
                    caso.lastGestionDate = self.sacConnector.getLastGestionControl(idMapsa=caso.idMapsa)
                    deudor.casoAsociado = caso
                    self.sacConnector.setMapsaCasoState(idMapsa=caso.idMapsa, newState=ACTIVO.lower().capitalize())
                    gestion: Gestion = Gestion(idJuicio=caso.idMapsa, timestamp=datetime.now(), tipo=PLPBREACHED, rutDeudor=caso.rutDeudor, nombreDeudor=caso.nombreDeudor + ' ' + caso.apellidoDeudor)
                    self.sacConnector.addGestion(gestion=gestion)
                    deudor.casoAsociado.nombreEstado = ACTIVO.lower().capitalize()
                mappedDeudores.append(deudor)
            plpBreachedRequest.deudores = mappedDeudores
            processedPLPBreachedRequests.append(plpBreachedRequest)
        return processedPLPBreachedRequests
    
    def getRequestResults(self) -> tuple:
        data: pd.DataFrame = pd.read_excel(PLPREQUESTSPATH)
        plpRequests: list[UnMappedRequest] = []
        plpBreachedRequests: list[UnMappedRequest] = []
        judicialCollectionRequests: list[UnMappedRequest] = []
        totalPLPRequests: int = 0
        totalPLPBreachedRequests: int = 0
        totalJCRequests: int = 0
        for index, row in data.iterrows():
            emisor = row['Emisor']
            asunto = row['Asunto']
            idMapsa = row['ID Mapsa']
            tipoSolicitud = row['Tipo']
            nombreDeudor = row['Deudor']
            rutDeudor = row['RUT Deudor']
            if tipoSolicitud == SOLICITUDPLP:
                totalPLPRequests += 1
                if idMapsa == 'No encontrado':
                    unmappedRequest: UnMappedRequest = UnMappedRequest(tipoSolicitud=SOLICITUDPLP,
                                                                    emisor=emisor,
                                                                    asunto=asunto,
                                                                    nombreDeudor=nombreDeudor,
                                                                    rutDeudor=rutDeudor)
                    plpRequests.append(unmappedRequest)
            elif tipoSolicitud == PLPINCUMPLIDO:
                totalPLPBreachedRequests += 1
                if idMapsa == 'No encontrado':
                    unmappedRequest: UnMappedRequest = UnMappedRequest(tipoSolicitud=SOLICITUDPLP,
                                                                    emisor=emisor,
                                                                    asunto=asunto,
                                                                    nombreDeudor=nombreDeudor,
                                                                    rutDeudor=rutDeudor)
                    plpBreachedRequests.append(unmappedRequest)
                    
            else:
                totalJCRequests += 1
                if idMapsa == 'No encontrado':
                    unmappedRequest: UnMappedRequest = UnMappedRequest(tipoSolicitud=tipoSolicitud,
                                                                       emisor=emisor,
                                                                       asunto=asunto,
                                                                       nombreDeudor=nombreDeudor,
                                                                       rutDeudor=rutDeudor)
                    judicialCollectionRequests.append(unmappedRequest)
                
        return (plpRequests, plpBreachedRequests, judicialCollectionRequests, totalPLPRequests, totalPLPBreachedRequests, totalJCRequests)
    
    def isRequestFileEmpty(self) -> bool:
        if not os.path.exists(PLPREQUESTSPATH):
            return True
        data = pd.read_excel(PLPREQUESTSPATH)
        numRows: int = data.shape[0]
        return not bool(numRows)
    
    def generateSummary(self, date = datetime.today()) -> str:
        data: tuple = self.getRequestResults()
        recurrentCasos: list = self.getRecurrentCasos()
        plpRequests: list[UnMappedRequest] = data[0]
        plpBreachedRequests: list[UnMappedRequest] = data[1]
        judicialCollectionRequests: list[UnMappedRequest] = data[2]
        totalPLPRequests: int = data[3]
        totalPLPBreachedRequests: int = data[4]
        totalJCRequests: int = data[5]
        totalRequests: int = totalPLPRequests + totalPLPBreachedRequests + totalJCRequests
        totalUnmappedRequests: int = len(plpRequests) + len(plpBreachedRequests) + len(judicialCollectionRequests)
        plpUnmappedRequests: int = len(plpRequests)
        plpBreachedUnmappedRequests: int = len(plpBreachedRequests)
        jcUnmappedRequests: int = len(judicialCollectionRequests)
        if not totalRequests:
            text: str = f'Buenas noches, \nAl día {transformDateToSpanish(date)} no se recibieron solicitudes.'
        if totalRequests:
            text: str = f'Buenas noches, \nAl día {transformDateToSpanish(date)} se recibieron {totalPLPRequests} solicitudes de PLP, {totalPLPBreachedRequests} PLPs incumplidos y {totalJCRequests} solicitudes de cobranza judicial.\n\n'
        if not totalUnmappedRequests:
            text += 'Todas las solicitudes fueron mapeadas y procesadas correctamente.'
        if plpRequests:
            text += f'{plpUnmappedRequests} solicitudes de PLP no pudieron asociarse a un caso: \n\n'
            text += '\n'.join([f'{index + 1}) {plpRequest.emisor} - {plpRequest.asunto} - {plpRequest.rutDeudor}' for index, plpRequest in enumerate(plpRequests)])
            text += '\n'
        text += '\n\n'
        if plpBreachedRequests:
            text += f'{plpBreachedUnmappedRequests} PLPs incumplidos no pudieron asociarse a un caso: \n\n'
            text += '\n'.join([f'{index + 1}) {plpBreachedRequest.emisor} - {plpBreachedRequest.asunto} - {plpBreachedRequest.nombreDeudor}' for index, plpBreachedRequest in enumerate(plpBreachedRequests)])
            text += '\n'
        text += '\n\n'
        if judicialCollectionRequests:
            text += f'{jcUnmappedRequests} solicitudes de cobranza judicial no pudieron asociarse a un caso: \n\n'
            text += '\n'.join([f'{index + 1}) {jcRequest.emisor} - {jcRequest.asunto} - {jcRequest.nombreDeudor}' for index, jcRequest in enumerate(judicialCollectionRequests)])
            text += '\n'
        text += '\n\n'
        if recurrentCasos:
            text += f'ATENCIÓN: {len(recurrentCasos)} casos han sido modificados en menos de {MIN_REQUEST_REPETITION_DELAY} días, cuyas gestiones fueron marcadas con amarillo: \n\n'
            text += '\n'.join([f'{index + 1}) {recurrentCaso[7]} - {recurrentCaso[6]} - {recurrentCaso[5]}' for index, recurrentCaso in enumerate(recurrentCasos)])
            text += '\n\n'
        text += 'Saludos, SACAutomations'
        return text
    
    def getRecurrentGestionesFromDB(self) -> list[Gestion]:
        data: pd.DataFrame = pd.read_excel(PLPREQUESTSPATH)
        idsCasos: list[int] = []
        for index, row in data.iterrows():
            idMapsa: str = str(row['ID Mapsa'])
            if idMapsa.isdigit():
                idsCasos.append(int(idMapsa))
        recurrentGestiones: list[Gestion] = self.sacConnector.getRecurrentGestiones(delay=MIN_REQUEST_REPETITION_DELAY, idsCasos=idsCasos)
        return recurrentGestiones
    
    def getRecurrentCasos(self) -> list:
        workbook = openpyxl.load_workbook(PLPREQUESTSPATH)
        worksheet = workbook.active
        rowsColored = []
        for row in worksheet.iter_rows():
            rowColored = False
            for cell in row:
                if cell.fill is not None and cell.fill.start_color.rgb == 'FFFFFF00':
                    rowColored = True
                    break
            if rowColored:
                rowData = [cell.value for cell in row]
                rowsColored.append(rowData)
        return rowsColored
    
    
        







        
        
