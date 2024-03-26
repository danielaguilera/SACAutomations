from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import shutil
import os
from tkinter import filedialog
from Clases.Boleta import Boleta
from Clases.PDFGenerator import PDFGenerator
from Clases.Destinatario import Destinatario
from Clases.Resumen import Resumen
from Utils.Metadata import *
from Utils.GlobalFunctions import *
from PyPDF2 import PdfReader, PdfMerger, PdfWriter
from Clases.SACConnector import SACConnector
from Clases.Cliente import Cliente
from Clases.AddServicioGUI import AddServicioGUI
from Clases.Servicio import Servicio
from Clases.Beneficiario import Beneficiario
from Clases.Caso import Caso
from Clases.ReportManager import ReportManager
from Clases.SACSenderJob import SACSenderJob
from datetime import date
from unidecode import unidecode
import fitz
import re
import time
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Border, Side, Alignment, NamedStyle


class App:
    def __init__(self):
        
        self.sacConnector: SACConnector = SACConnector()
        self.clientes: list[Cliente] = self.sacConnector.getAllClientes()
        self.beneficiarios: list[Beneficiario] = self.sacConnector.getAllBeneficiarios()
        self.destinatarios: list[Destinatario] = self.sacConnector.getAllDestinatarios()
        self.casos: list[Caso] = []
        
        self.master = Tk()
        self.master.title("SAC App")
        self.master.protocol("WM_DELETE_WINDOW", self.onClosingWindow)
        
        self.thumbnailFrame = Frame(master=self.master)
        self.thumbnailFrame.pack(side=RIGHT)
        self.boletaImage = None
        self.thumbnail = Label(master=self.thumbnailFrame)
        
        self.stateFrame = LabelFrame(master=self.master)
        self.stateFrame.pack(expand=True, fill=BOTH)
        
        Label(master=self.stateFrame, text='Datos generales: ', font=('Helvetica bold', 10, 'bold')).pack(side=TOP)
        
        self.uploadedBoletaFrame = Frame(master=self.stateFrame)
        self.uploadedBoletaFrame.pack(expand=True, fill=BOTH)
        self.uploadedBoletaLabel = Label(master=self.uploadedBoletaFrame, font=('Helvetica bold', 10), text='No se ha subido boleta')
        self.uploadedBoletaLabel.pack(side=LEFT)
        self.boletaUploadButton = Button(self.uploadedBoletaFrame, text="Subir boleta/factura", fg = 'black', bg='RoyalBlue1', command=self.selectBoletaPDF)
        self.boletaUploadButton.pack(side=LEFT, padx=5)
        self.boletaResetButton = Button(self.uploadedBoletaFrame, text="Quitar boleta", fg = 'black', bg='indian red', command=self.clearForm)
        self.boletaResetButton.pack(side=LEFT, padx=5)
        
        self.uploadedAnexosFrame = Frame(master=self.stateFrame)
        self.uploadedAnexosFrame.pack(expand=True, fill=BOTH)
        self.uploadedAnexosLabel = Label(master=self.uploadedAnexosFrame, font=('Helvetica bold', 10), text='No se han subido anexos')
        self.uploadedAnexosLabel.pack(side=LEFT)
        self.anexoUploadButton = Button(self.uploadedAnexosFrame, text="Subir anexo", fg = 'black', bg='RoyalBlue1', command=self.selectAnexoPDF)
        self.anexoUploadButton.pack(side=LEFT, padx=5)
        self.anexosResetButton = Button(self.uploadedAnexosFrame, text="Quitar anexos", fg = 'black', bg='indian red', command=self.resetAnexos)
        self.anexosResetButton.pack(side=LEFT, padx=5)
        
        self.numBoletaFrame = Frame(master=self.stateFrame)
        self.numBoletaFrame.pack(expand=True, fill=BOTH)
        self.numBoletaLabel = Label(master=self.numBoletaFrame, text='N° Boleta')
        self.numBoletaLabel.pack(side=LEFT)
        self.numBoletaEntry = Entry(master=self.numBoletaFrame, justify='center')
        self.numBoletaEntry.pack(side=LEFT, padx=5)
        
        self.fechaBoletaFrame = Frame(master=self.stateFrame)
        self.fechaBoletaFrame.pack(expand=True, fill=BOTH)
        self.fechaBoletaLabel = Label(master=self.fechaBoletaFrame, text='Fecha emisión (dd-mm-AAAA)')
        self.fechaBoletaLabel.pack(side=LEFT) 
        self.fechaBoletaEntry: Entry = Entry(master=self.fechaBoletaFrame, justify='center')
        self.fechaBoletaEntry.pack(side=LEFT, padx=5)
        
        self.rutBeneficiarioFrame = Frame(master=self.stateFrame)
        self.rutBeneficiarioFrame.pack(expand=True, fill=BOTH)
        self.rutBeneficiarioLabel = Label(master=self.rutBeneficiarioFrame, text='RUT Beneficiario')
        self.rutBeneficiarioLabel.pack(side=LEFT) 
        self.rutBeneficiarioEntry = Entry(master=self.rutBeneficiarioFrame, justify='center')
        self.rutBeneficiarioEntry.pack(side=LEFT, padx=5)
        self.rutBeneficiarioEntry.bind("<KeyRelease>", self.findBeneficiario)
        
        self.nombreBeneficiarioFrame = Frame(master=self.stateFrame)
        self.nombreBeneficiarioFrame.pack(expand=True, fill=BOTH)
        self.nombreBeneficiarioLabel = Label(master=self.nombreBeneficiarioFrame, text='Nombre o razón social')
        self.nombreBeneficiarioLabel.pack(side=LEFT)
        self.nombreBeneficiarioDropdown = ttk.Combobox(master=self.nombreBeneficiarioFrame, state='readonly', values=[beneficiario.nombreBeneficiario for beneficiario in self.beneficiarios], width=40)
        self.nombreBeneficiarioDropdown.pack(side=LEFT, padx=5)
        self.nombreBeneficiarioDropdown.bind("<<ComboboxSelected>>", self.assignBeneficiario)
        
        self.rutDeudorFrame = Frame(master=self.stateFrame)
        self.rutDeudorFrame.pack(expand=True, fill=BOTH)
        self.rutDeudorLabel = Label(master=self.rutDeudorFrame, text='RUT Deudor')
        self.rutDeudorLabel.pack(side=LEFT) 
        self.rutDeudorEntry = Entry(master=self.rutDeudorFrame, justify='center')
        self.rutDeudorEntry.pack(side=LEFT, padx=5)
        self.rutDeudorEntry.bind("<KeyRelease>", self.populateCasos) 
        
        self.nombreDeudorFrame = Frame(master=self.stateFrame)
        self.nombreDeudorFrame.pack(expand=True, fill=BOTH)
        self.nombreDeudorLabel = Label(master=self.nombreDeudorFrame, text='Nombre deudor')
        self.nombreDeudorLabel.pack(side=LEFT) 
        self.nombreDeudorEntry = Entry(master=self.nombreDeudorFrame, justify='center')
        self.nombreDeudorEntry.pack(side=LEFT, padx=5)
        self.nombreDeudorEntry.bind("<KeyRelease>", self.populateCasos) 

        self.apellidoDeudorFrame = Frame(master=self.stateFrame)
        self.apellidoDeudorFrame.pack(expand=True, fill=BOTH)
        self.apellidoDeudorLabel = Label(master=self.apellidoDeudorFrame, text='Apellido deudor')
        self.apellidoDeudorLabel.pack(side=LEFT) 
        self.apellidoDeudorEntry = Entry(master=self.apellidoDeudorFrame, justify='center')
        self.apellidoDeudorEntry.pack(side=LEFT, padx=5)
        self.apellidoDeudorEntry.bind("<KeyRelease>", self.populateCasos)
        
        self.clienteFrame = Frame(master=self.stateFrame)
        self.clienteFrame.pack(expand=True, fill=BOTH)
        self.clienteLabel = Label(master=self.clienteFrame, text='Cliente')
        self.clienteLabel.pack(side=LEFT)
        self.clienteDropdown = ttk.Combobox(master=self.clienteFrame, state='readonly', values=[cliente.nombreCliente for cliente in self.clientes] + ['Ninguno'], justify='center')
        self.clienteDropdown.pack(side=LEFT, padx=5)
        self.clienteDropdown.bind("<<ComboboxSelected>>", self.clienteSelectionEvent)
        self.rendicionLabel = Label(master=self.clienteFrame, text='# Rendición')
        self.rendicionLabel.pack(side=LEFT, padx=5)
        self.rendicionEntry = Entry(master=self.clienteFrame, justify='center')
        self.rendicionEntry.pack(side=LEFT, padx=5)
        self.rendicionEntry.insert(0, '-')
        
        self.gastoTotalFrame = Frame(master=self.stateFrame)
        self.gastoTotalFrame.pack(expand=True, fill=BOTH)
        self.gastoTotalLabel = Label(master=self.gastoTotalFrame, text='Total ($)')
        self.gastoTotalLabel.pack(side=LEFT)
        self.gastoTotalEntry = Entry(master=self.gastoTotalFrame, justify='center')
        self.gastoTotalEntry.pack(side=LEFT, padx=5)
        
        self.destinatarioFrame = Frame(master=self.stateFrame)
        self.destinatarioFrame.pack(expand=True, fill=BOTH)
        self.destinatarioLabel = Label(master=self.destinatarioFrame, text='Se envía a: ')
        self.destinatarioLabel.pack(side=LEFT)

        self.saveFrame = Frame(master=self.stateFrame)
        self.saveFrame.pack(expand=True, fill=BOTH)
        Label(master=self.saveFrame, text='Acciones').pack(side=LEFT)
        self.saveButton = Button(self.saveFrame, text='Generar reporte y Guardar', font=('Helvetica bold', 10), fg = 'black', bg='RoyalBlue1', command=self.saveChanges)
        self.saveButton.pack(side=LEFT, padx=10)
        self.clearFormButton = Button(self.saveFrame, text='Borrar formulario', font=('Helvetica bold', 10), fg = 'black', bg='indian red', command=self.beginClearForm)
        self.clearFormButton.pack(side=LEFT, padx=10)
        self.manageReportsButton = Button(self.saveFrame, text='Ver reportes a enviar', font=('Helvetica bold', 10),  fg = 'black', bg='lawngreen', command=self.runReportManager)
        self.manageReportsButton.pack(side=LEFT, padx=10)
        
        self.casosFrame = Frame(master=self.master)
        self.casosFrame.pack(expand=True, fill=BOTH)
        Label(master=self.casosFrame, text='Asociar caso a boleta: ', font=('Helvetica bold', 10, 'bold')).pack(side=TOP)
        self.casosColumns = ['ID Mapsa', 'Estado', 'Fecha Asignación', 'Bsecs', 'RUT Deudor', 'Apellido Deudor', 'Nombre Deudor', 'Cliente', 'Facturar a']
        self.casosTable = ttk.Treeview(master=self.casosFrame, columns=self.casosColumns, show='headings', height=3)
        displayColumns: list[str] = []
        for heading in self.casosColumns:
            self.casosTable.heading(heading, text=heading)
            self.casosTable.column(heading)
            if heading not in ['Estado', 'Bsecs', 'Cliente']:
                displayColumns.append(heading)
        self.casosTable["displaycolumns"] = displayColumns
        self.casosTable.pack(expand=True, fill=BOTH, anchor=CENTER)
        self.casosTable.bind('<<TreeviewSelect>>', self.selectCaso)
        
        self.serviciosFrame = Frame(master=self.master)
        self.serviciosFrame.pack(expand=True, fill=BOTH)
        Label(master=self.serviciosFrame, text='Asociar servicios a boleta: ', font=('Helvetica bold', 10, 'bold')).pack(side=TOP)
        self.addServicioButton = Button(master=self.serviciosFrame, text='Agregar servicio', fg = 'black', bg='SeaGreen3', command=self.openServicioGUI)
        self.addServicioButton.pack(side=LEFT, padx=5)
        self.deleteServicioButton = Button(master=self.serviciosFrame, text='Eliminar servicio', fg = 'black', bg='indian red', command=self.removeServicio)
        self.deleteServicioButton.pack(side=LEFT, padx=5)
        self.serviciosColumns = ['Código', 'Nota', 'Monto']
        self.serviciosTable = ttk.Treeview(master=self.serviciosFrame, columns=self.serviciosColumns, show='headings', height=3)
        self.serviciosTable.heading('Código', text='Código')
        self.serviciosTable.heading('Nota', text='Nota')
        self.serviciosTable.heading('Monto', text='Monto ($)')
        self.serviciosTable.pack(expand=True, fill=BOTH, anchor=CENTER)
        self.serviciosTable.insert('', END, values=('TOTAL', '-', 0), iid='total')
        
        self.boletaPath: str = ''
        self.anexosPaths: list[str] = []
        self.destinatario: Destinatario = None
        self.cc: list[str] = []
        
        self.saveFrame = LabelFrame(master=self.master)
        self.saveFrame.pack(expand=True, fill=BOTH)
        
        self.master.mainloop()

    @property
    def numBoleta(self) -> int:
        userInput : str = self.numBoletaEntry.get()
        if userInput.isdigit():
            return int(userInput)
        else:
            return 0
        
    @property
    def idMapsa(self) -> int:
        if not self.casosTable.focus():
            return 0
        return int(self.casosTable.item(self.casosTable.focus())['values'][0])

    @property
    def numAnexos(self) -> int:
        return len(self.anexosPaths)
    
    @property
    def servicioSum(self) -> int:
        total: int = 0
        for iid in self.serviciosTable.get_children():
            if iid == 'total':
                continue
            monto: int = self.serviciosTable.item(iid)['values'][2]
            total += int(monto)
        return total
    
    @property
    def addedServicios(self) -> int:
        return len(self.serviciosTable.get_children())
    
    def getClienteRendicion(self) -> int:
        idMapsa: int = self.casosTable.item(self.casosTable.focus())['values'][0]
        casoSet: Caso | None = next((caso for caso in self.casos if caso.idMapsa == idMapsa), None)
        if not casoSet:
            return 0
        if not self.destinatario:
            return 0
        nombreDestinatario = self.destinatario.nombreDestinatario
        nombreCliente = casoSet.nombreCliente
        excelMatrixRoot = f'{DELIVEREDDATAPATH}/{nombreDestinatario}/Rendición {nombreCliente}.xlsx'
        if not os.path.exists(excelMatrixRoot):
            return 0
        workbook = openpyxl.load_workbook(excelMatrixRoot)
        sheet = workbook.active
        cell = sheet.cell(row = 1, column = 8)
        label: str = cell.value
        return int(label.split(' ')[2].split('/')[0])
    
    def runReportManager(self):
        if not os.path.exists(DELIVEREDDATAPATH):
            messagebox.showerror(title='ERROR', message='No se han cargado boletas para enviar.\nNo hay nada que mostrar.')
            return
        reportManager: ReportManager = ReportManager(container=self)

    def rootFilesOpen(self) -> bool:
        try:
            os.listdir(SAC_PATH)
            return False
        except OSError as e:
            if e.errno == 13:
                messagebox.showerror(title='Error', message='Por favor cierra la carpeta SAC \n antes de continuar')  
                return True 
            else:
                raise e
        
    def validData(self) -> bool:
        if not self.boletaPath:
            messagebox.showerror(title='Error', message='Falta agregar la boleta/factura')
            return False
        if not self.numBoletaEntry.get().isdigit():
            messagebox.showerror(title='Error', message='El número de boleta debe ser un número entero')
            return False
        if not self.rutBeneficiarioEntry.get() in [beneficiario.rutBeneficiario for beneficiario in self.beneficiarios]:
            messagebox.showerror(title='Error', message='No existe beneficiario con ese rut')
            return False
        if not self.nombreBeneficiarioDropdown.get():
            messagebox.showerror(title='Error', message='Beneficiario no identificado')
            return False
        if not self.casosTable.focus():
            messagebox.showerror(title='Error', message='Debes seleccionar un caso')
            return False
        if not self.serviciosTable.get_children():
            messagebox.showerror(title='Error', message='Debes agregar al menos 1 servicio en la boleta')
            return False
        if not self.gastoTotalEntry.get().isdigit():
            messagebox.showerror(title='Error', message='El total de la boleta debe ser un número entero')
            return False
        if int(self.gastoTotalEntry.get()) != self.servicioSum:
            messagebox.showerror(title='Error', message='El total de la boleta debe ser igual a la suma de los montos de los servicios')
            return False
        if not self.destinatario:
            messagebox.showerror(title='Error', message='Falta escoger un destinatario')
            return False
        if not validDateFormat(self.fechaBoletaEntry.get()):
            messagebox.showerror(title='Error', message='Debes poner la fecha en formato dd-mm-AAAA')
            return False
        if not validRendicionNumber(self.rendicionEntry.get()):
            messagebox.showerror(title='Error', message='El número de rendición debe ser un entero')
            return False
        if self.casoAlreadyInBoleta():
            messagebox.showerror(title='Error', message='Ya hay una boleta asociada a este caso')
            return False
        if self.boletaAlreadyGenerated():
            messagebox.showerror(title='Error', message=f'Ya existe un reporte para la boleta # {self.numBoletaEntry.get()}')
            return False
        return True     
    
    def casoAlreadyInBoleta(self) -> bool:
        idMapsaSelected: int = int(self.casosTable.item(self.casosTable.focus())['values'][0])
        return bool(self.sacConnector.getBoletasFromCaso(idCaso=idMapsaSelected))
    
    def boletaAlreadyGenerated(self):
        if not self.numBoletaEntry.get().isdigit():
            return False
        numBoleta: int = int(self.numBoletaEntry.get())
        return bool(self.sacConnector.getBoletaData(numBoleta=numBoleta))
                
    def saveChanges(self):
        try:
            start = time.process_time() 
            if not self.validData():
                return  
            if not messagebox.askyesno(title='Aviso', message='¿Estás segur@ de querer guardar los datos?'):
                return
            dataSelected: list = self.casosTable.item(self.casosTable.focus())['values']
            idMapsa: int = dataSelected[0]
            numBoleta: int = int(self.numBoletaEntry.get())
            fechaEmision: date = datetime.strptime(self.fechaBoletaEntry.get(), '%d-%m-%Y').date()
            rutBeneficiario: str = self.rutBeneficiarioEntry.get()
            servicios: list[Servicio] = []
            for iid in self.serviciosTable.get_children():
                if iid == 'total':
                    continue
                codigo, nota, monto = self.serviciosTable.item(iid)['values']
                codigo = codigo.replace(': ','')
                servicios.append(Servicio(codigo=codigo, nota=nota, monto=monto))
            boleta: Boleta = Boleta(idMapsa=idMapsa, 
                                    numBoleta=numBoleta, 
                                    fechaEmision=fechaEmision, 
                                    rutBeneficiario=rutBeneficiario, 
                                    servicios=servicios)
            self.sacConnector.insertBoletaData(boleta=boleta)        
            numBoleta: int = int(self.numBoletaEntry.get())
            idMapsa: int = boleta.idMapsa
            if not os.path.exists(DELIVEREDDATAPATH):
                os.makedirs(DELIVEREDDATAPATH)
            if not os.path.exists(f'{DELIVEREDDATAPATH}/{self.destinatario.nombreDestinatario}/{numBoleta}_{idMapsa}'):
                os.makedirs(f'{DELIVEREDDATAPATH}/{self.destinatario.nombreDestinatario}/{numBoleta}_{idMapsa}')
            merger: PdfMerger = PdfMerger()
            for root in self.anexosPaths:
                merger.append(root)
            if self.anexosPaths:
                merger.write(f'{DELIVEREDDATAPATH}/{self.destinatario.nombreDestinatario}/{numBoleta}_{idMapsa}/Anexo_{numBoleta}.pdf')
            merger.close()
            reader = PdfReader(self.boletaPath, strict=False)
            writer = PdfWriter()
            page = reader.pages[0]
            writer.add_page(page)
            with open(f'{DELIVEREDDATAPATH}/{self.destinatario.nombreDestinatario}/{numBoleta}_{idMapsa}/Boleta_{numBoleta}.pdf', 'wb') as file:
                writer.write(file)
            self.generateReport(boleta=boleta)
            self.saveParams()
            print('Tiempo en subir boleta: ' + str(time.process_time() - start))
            if self.checkBoletaInDB() and self.checkBoletainFile():
                messagebox.showinfo(title='Mensaje', message=f'Boleta n°{numBoleta} ingresada exitosamente')
                with open(ACTIVITYLOGFILE, 'a') as file:
                    file.write(f'{str(datetime.now())}: {USER} añadió boleta a enviar (NUMERO BOLETA: {numBoleta} - ID MAPSA: {idMapsa}) para {self.destinatario.nombreDestinatario}\n')
                    self.clearForm()
            else:
                messagebox.showinfo(title='Error', message=f'Boleta n°{numBoleta} no se pudo subir correctamente. Por favor intentar nuevamente')
                self.clearBoletaFiles()
                self.clearBoletaFromDB()
        except Exception as e:
           messagebox.showinfo(title='Error', message=str(e))
           self.clearBoletaFiles()
           self.clearBoletaFromDB()

    def checkBoletainFile(self) -> bool:
        boletaExists: bool = os.path.exists(f'{DELIVEREDDATAPATH}/{self.destinatario.nombreDestinatario}/{self.numBoleta}_{self.idMapsa}/Boleta_{self.numBoleta}.pdf')
        reporteExists: bool = os.path.exists(f'{DELIVEREDDATAPATH}/{self.destinatario.nombreDestinatario}/{self.numBoleta}_{self.idMapsa}/Reporte_{self.numBoleta}.pdf')
        dataExists: bool = os.path.exists(f'{DELIVEREDDATAPATH}/{self.destinatario.nombreDestinatario}/{self.numBoleta}_{self.idMapsa}/Data_{self.numBoleta}.txt')
        return boletaExists and reporteExists and dataExists

    def checkBoletaInDB(self) -> bool:
        return len(self.serviciosTable.get_children("")) - 1 == len(self.sacConnector.getBoletaServicios(numBoleta=self.numBoleta, idMapsa=self.idMapsa))

    def clearBoletaFromDB(self):
        self.sacConnector.deleteBoletaData(numBoleta=self.numBoleta, idMapsa=self.idMapsa)

    def clearBoletaFiles(self):
        if os.path.exists(f'{DELIVEREDDATAPATH}/{self.destinatario.nombreDestinatario}/{self.numBoleta}_{self.idMapsa}'):
            shutil.rmtree(f'{DELIVEREDDATAPATH}/{self.destinatario.nombreDestinatario}/{self.numBoleta}_{self.idMapsa}')
        if not os.listdir(f'{DELIVEREDDATAPATH}/{self.destinatario.nombreDestinatario}'):
            shutil.rmtree(f'{DELIVEREDDATAPATH}/{self.destinatario.nombreDestinatario}')
        if not os.listdir(f'{DELIVEREDDATAPATH}'):
            shutil.rmtree(f'{DELIVEREDDATAPATH}')
        
    def generateUnifiedDocument(self):
        for nombreDestinatario in os.listdir(path=f'{DELIVEREDDATAPATH}'):
            pdfMerger: PdfMerger = PdfMerger()
            for path in os.listdir(path=f'{DELIVEREDDATAPATH}/{nombreDestinatario}'):
                if path != 'Documento.pdf':
                    numBoleta, idMapsa = (int(x) for x in path.strip().split('_'))
                    reportePath: str = f'{DELIVEREDDATAPATH}/{nombreDestinatario}/{path}/Reporte_{numBoleta}.pdf'
                    boletaPath: str = f'{DELIVEREDDATAPATH}/{nombreDestinatario}/{path}/Boleta_{numBoleta}.pdf'
                    anexoPath: str = f'{DELIVEREDDATAPATH}/{nombreDestinatario}/{path}/Anexo_{numBoleta}.pdf'
                    pdfMerger.append(reportePath)
                    pdfMerger.append(boletaPath)
                    if os.path.exists(anexoPath):
                        pdfMerger.append(anexoPath)
            pdfMerger.write(f'{DELIVEREDDATAPATH}/{nombreDestinatario}/Documento.pdf')
        pdfMerger.close()

    def updateUnifiedDocument(self):
        nombreDestinatario: str = self.destinatario.nombreDestinatario
        numBoleta: int = int(self.numBoletaEntry.get())
        idMapsa: int = int(self.casosTable.item(self.casosTable.focus())['values'][0])
        pdfMerger: PdfMerger = PdfMerger()
        reportePath: str = f'{DELIVEREDDATAPATH}/{nombreDestinatario}/{numBoleta}_{idMapsa}/Reporte_{numBoleta}.pdf'
        boletaPath: str = f'{DELIVEREDDATAPATH}/{nombreDestinatario}/{numBoleta}_{idMapsa}/Boleta_{numBoleta}.pdf'
        anexoPath: str = f'{DELIVEREDDATAPATH}/{nombreDestinatario}/{numBoleta}_{idMapsa}/Anexo_{numBoleta}.pdf'
        if os.path.exists(f'{DELIVEREDDATAPATH}/{nombreDestinatario}/Documento.pdf'):
            os.rename(f'{DELIVEREDDATAPATH}/{nombreDestinatario}/Documento.pdf', f'{DELIVEREDDATAPATH}/{nombreDestinatario}/Documento_ant.pdf')
            pdfMerger.append(f'{DELIVEREDDATAPATH}/{nombreDestinatario}/Documento_ant.pdf')
        pdfMerger.append(reportePath)
        pdfMerger.append(boletaPath)
        if os.path.exists(anexoPath):
            pdfMerger.append(anexoPath)
        pdfMerger.write(f'{DELIVEREDDATAPATH}/{nombreDestinatario}/Documento.pdf')
        pdfMerger.close()
        if os.path.exists(f'{DELIVEREDDATAPATH}/{nombreDestinatario}/Documento_ant.pdf'):
            os.remove(f'{DELIVEREDDATAPATH}/{nombreDestinatario}/Documento_ant.pdf')
        
    def saveDeudorName(self):
        idMapsaSet: int = self.casosTable.item(self.casosTable.focus())['values'][0]
        numBoletaSet: int = int(self.numBoletaEntry.get())
        with open(f'{DELIVEREDDATAPATH}/{self.destinatario.nombreDestinatario}/{numBoletaSet}_{idMapsaSet}/DeudorName.txt', 'w') as file:
            file.write(self.nombreDeudorEntry.get())

    def saveCC(self):
        if not self.cc:
            return
        idMapsaSet: int = self.casosTable.item(self.casosTable.focus())['values'][0]
        numBoletaSet: int = int(self.numBoletaEntry.get())
        with open(f'{DELIVEREDDATAPATH}/{self.destinatario.nombreDestinatario}/{numBoletaSet}_{idMapsaSet}/CC.txt', 'w') as file:
            file.write(','.join(self.cc))
    
    def saveParams(self):
        idMapsaSet: int = self.casosTable.item(self.casosTable.focus())['values'][0]
        numBoletaSet: int = int(self.numBoletaEntry.get())
        destinatarioSet: Destinatario = self.destinatario
        beneficiarioSet: str = self.nombreBeneficiarioDropdown.get()
        clienteSet: str = self.clienteDropdown.get()
        deudorSet: str = self.apellidoDeudorEntry.get()
        montoTotalSet: str = self.gastoTotalEntry.get()     
        with open(f'{DELIVEREDDATAPATH}/{destinatarioSet.nombreDestinatario}/{numBoletaSet}_{idMapsaSet}/Data_{numBoletaSet}.txt', 'w') as file:
            file.write(f'{destinatarioSet.nombreDestinatario},{destinatarioSet.correoDestinatario},{numBoletaSet},{idMapsaSet},{beneficiarioSet},{clienteSet},{deudorSet},{montoTotalSet}')

    def generateReport(self, boleta: Boleta):
        idMapsa: int = self.casosTable.item(self.casosTable.focus())['values'][0]
        casoSet: Caso | None = next((caso for caso in self.casos if caso.idMapsa == idMapsa), None)
        destinatarioSet: Destinatario = self.destinatario
        serviciosSet: list[Servicio] = boleta.servicios
        numeroRendicion: int = int(self.rendicionEntry.get())
        beneficiarioSet: Beneficiario = Beneficiario(rutBeneficiario=self.rutBeneficiarioEntry.get(), 
                                                     nombreBeneficiario=self.nombreBeneficiarioDropdown.get())
        clienteSet: Cliente = next((cliente for cliente in self.clientes if cliente.idCliente == casoSet.idCliente), None)
        resumenBoleta: Resumen = Resumen(boleta=boleta, caso=casoSet, 
                                         destinatario=destinatarioSet, 
                                         servicios=serviciosSet, 
                                         beneficiario=beneficiarioSet, 
                                         cliente=clienteSet,
                                         numeroRendicion=numeroRendicion)
        pdfGenerator: PDFGenerator = PDFGenerator()
        pdfGenerator.generateReporte(resumenBoleta=resumenBoleta)
        
    def fileIsPDF(self, filePath: str):
        try:
            PdfReader(filePath, strict=False)
            return True
        except Exception:
            return False
    
    def resetBoleta(self):
        self.boletaPath = ''
        self.uploadedBoletaLabel.config(text='No se ha subido boleta')
        self.numBoletaEntry.delete(0, END)
        
    def resetAnexos(self):
        self.anexosPaths.clear()
        self.uploadedAnexosLabel.config(text='No se han subido anexos')
                
    def selectAnexoPDF(self):
        filePath = filedialog.askopenfilename(filetypes=[("PDF Files", "*.pdf")])
        if self.fileIsPDF(filePath):
            self.anexosPaths.append(filePath)
            messagebox.showinfo(title='Mensaje', message='Anexo agregado correctamente')
            self.uploadedAnexosLabel.config(text=f'Anexos subidos: {self.numAnexos}')
        else:
            messagebox.showerror(title='Error', message='Archivo no válido')
            
    def selectBoletaPDF(self):
        filePath = filedialog.askopenfilename(filetypes=[("PDF Files", "*.pdf")])
        if self.fileIsPDF(filePath):
            self.boletaPath = filePath
            messagebox.showinfo(title='Mensaje', message='Boleta subida correctamente')
            self.uploadedBoletaLabel.config(text=f'Boleta subida')
            reader: PdfReader = PdfReader(self.boletaPath, strict=False)
            text: str = reader.pages[0].extract_text()
            self.getNumeroBoletaFromFile(text=text)
            self.getRUTBeneficiarioFromFile(text=text)
            self.getFechaFromFile(text=text)
            self.getGastoTotalFromFile(text=text)
            self.displayThumbnail()
            if self.boletaAlreadyGenerated():
                messagebox.showerror(title='Error', message=f'Ya existe un reporte para la boleta # {self.numBoletaEntry.get()}')
                self.clearForm()
            return False
        else:
            messagebox.showerror(title='Error', message='Archivo no válido')
            
    def displayThumbnail(self):
        zoom_x = 1.0  # horizontal zoom
        zoom_y = 1.0  # vertical zoom
        mat = fitz.Matrix(zoom_x, zoom_y)  # zoom factor 2 in each dimension

        filename = self.boletaPath
        doc = fitz.open(filename)  # open document
        pix = doc[0].get_pixmap(matrix=mat)  # render page to an image
        pix.save("thumbnail.png")  # store image as a PNG
        doc.close()
            
        self.boletaImage = PhotoImage(file='thumbnail.png')
        self.thumbnailFrame.pack(side=RIGHT)
        self.thumbnail.config(image=self.boletaImage)
        self.thumbnail.pack()

    def fileIsFactura(self, text: str):
        text = text.replace('\n', ' ')
        if FACTURAKEYPHRASE in text:
            return True
        return False
    
    def getNumeroBoletaFromFile(self, text: str):
        try:
            if self.fileIsFactura(text=text):
                numberIndex: int = text.find('º')
            else:
                numberIndex: int = text.find('°')
            subString: str = text[numberIndex::]
            endIndex: int = subString.find('\n')
            subString = subString[0: endIndex + 1]
            numBoleta = extractNumberFromText(subString)
            if numBoleta.isdigit():
                self.numBoletaEntry.delete(0, END)
                self.numBoletaEntry.insert(0, numBoleta)
        except Exception:
            pass

    def getRUTBeneficiarioFromFile(self, text: str):
        try:
            text: str = unidecode(text.strip())
            if self.fileIsFactura(text=text):
                rutBeneficiario: str = re.findall(r"R\.U\.T\.:(.*)$", text, re.MULTILINE)[1].strip()
            else:
                rutBeneficiario: str = re.findall(r"RUT:(.*)$", text, re.MULTILINE)[0].strip()
            rutBeneficiario = correctRUTFormat(rutBeneficiario)
            self.rutBeneficiarioEntry.delete(0, END)
            self.rutBeneficiarioEntry.insert(0, rutBeneficiario)
            beneficiarioFound: Beneficiario = self.sacConnector.findBeneficiario(rutBeneficiario=rutBeneficiario)
            if beneficiarioFound:
                for index, beneficiario in enumerate(self.beneficiarios):
                    if beneficiario.nombreBeneficiario == beneficiarioFound.nombreBeneficiario:
                        self.nombreBeneficiarioDropdown.current(newindex=index)
                        return
            self.nombreBeneficiarioDropdown.set('Sin resultados')
        except Exception:
            pass
        
    def getFechaFromFile(self, text: str):
        try:
            text: str = unidecode(text.strip())
            if self.fileIsFactura(text=text):
                pattern: str = r'Fecha Emision: (.+)'
            else:
                pattern: str = r'Fecha: (.+)'
            patternMatch: re.match | None = re.search(pattern, text)
            fechaEmisionString: str = patternMatch.group(1) if patternMatch else ''
            fechaEmision: date = getDateFromSpanishFormat(stringDate=fechaEmisionString)
            self.fechaBoletaEntry.delete(0, END)
            self.fechaBoletaEntry.insert(0, fechaEmision.strftime("%d-%m-%Y"))
        except Exception:
            pass
        
    def getGastoTotalFromFile(self, text: str):
        try:
            text: str = unidecode(text.strip())
            if self.fileIsFactura(text=text):
                pattern: str = r'TOTAL \$(\d+(?:\.\d+)?)'
            else:
                pattern: str = r'Total Honorarios \$: (\d+(?:\.\d+)?)'
            patternMatch: re.match[str] | None = re.search(pattern, text)
            total: int | str = int(patternMatch.group(1).replace('.','')) if patternMatch else ''   
            self.gastoTotalEntry.delete(0, END)
            self.gastoTotalEntry.insert(0, total)
        except Exception:
            pass
        
    def runSender(self):
        if not messagebox.askyesno(title='Aviso', message='Se enviarán todas las boletas de esta semana. \nEsto puede tardar varios minutos. \n¿Deseas continuar?'):
            return
        try:
            if not os.path.exists(DELIVEREDDATAPATH):
                messagebox.showerror(title='Error', message='No hay reportes para enviar')
                return
            senderJob: SACSenderJob = SACSenderJob()
            senderJob.generateUnifiedDocument()
            senderJob.sendReports()
            messagebox.showinfo(title='Éxito', message='Reportes enviados')
            self.clearForm()
        except Exception as e:
            print(e)
            messagebox.showerror(title='Error', message='SAC Sender no pudo ejecutarse')
        
    def openServicioGUI(self):
        addServicioGUI: AddServicioGUI = AddServicioGUI(container=self)
        
    def removeServicio(self):
        if self.serviciosTable.focus() == 'total' or not self.serviciosTable.selection():
            return
        selectedItem = self.serviciosTable.selection()[0]
        self.serviciosTable.delete(selectedItem)
        self.serviciosTable.configure(height=self.addedServicios)
        self.serviciosTable.item('total', values = ('TOTAL', '-', self.servicioSum))
    
    def addServicio(self, servicio: Servicio):
        self.serviciosTable.insert('', END, values=(servicio.codigo, servicio.nota, servicio.monto))
        self.serviciosTable.item('total', values = ('TOTAL', '-', self.servicioSum))
        self.serviciosTable.configure(height=self.addedServicios)
        
    def clienteSelectionEvent(self, key=None):
        self.setDestinatario()
        self.populateCasos()    
    
    def setDestinatario(self, key=None):
        try:
            cliente: Cliente
            for cliente in self.clientes:
                if self.clienteDropdown.get() == cliente.nombreCliente:
                    idCliente: int = cliente.idCliente
            selectedDestinatario: Destinatario = self.sacConnector.getDestinatarioByCliente(idCliente=idCliente)
            self.cc: list[str] = selectedDestinatario.cc
            self.destinatario = selectedDestinatario
            text: str = f"Se envía a: {self.destinatario.nombreDestinatario} ({self.destinatario.correoDestinatario})"
            if self.cc:
                text += f" con copia a {', '.join(self.cc)}"
            self.destinatarioLabel.config(text=text)
        except Exception:
            pass
    
    def populateCasos(self, key=None):
        if len(self.casosTable.selection()) > 0:
            self.casosTable.selection_remove(self.casosTable.selection()[0])
        rutDeudor: str = self.rutDeudorEntry.get()
        selectedClienteName: str = self.clienteDropdown.get()
        idCliente: int | None = None
        if selectedClienteName:
            cliente: Cliente
            for cliente in self.clientes:
                if cliente.nombreCliente == selectedClienteName:
                    idCliente = cliente.idCliente
                    break
        apellidoDeudor: str = self.apellidoDeudorEntry.get()
        nombreDeudor: str = self.nombreDeudorEntry.get()
        self.casos = self.sacConnector.getPossibleMapsaCasos(rutDeudor=rutDeudor, idCliente=idCliente, apellidoDeudor=apellidoDeudor, nombreDeudor=nombreDeudor)
        caso: Caso
        self.casosTable.delete(*self.casosTable.get_children())
        for caso in self.casos:
            self.casosTable.insert('', END, values=(caso.idMapsa, caso.nombreEstado, caso.fechaAsignado.strftime('%d-%m-%Y') if caso.fechaAsignado else '', caso.bsecs, caso.rutDeudor, caso.apellidoDeudor, caso.nombreDeudor, caso.nombreCliente, caso.factura))

    def assignBeneficiario(self, key=None):
        nombreBeneficiario: str = self.nombreBeneficiarioDropdown.get()
        beneficiario: Beneficiario
        for beneficiario in self.beneficiarios:
            if beneficiario.nombreBeneficiario == nombreBeneficiario:
                self.rutBeneficiarioEntry.delete(0, END)
                self.rutBeneficiarioEntry.insert(0, beneficiario.rutBeneficiario)
                return
            
    def findBeneficiario(self, key=None):
        rutBeneficiario: str = self.rutBeneficiarioEntry.get()
        beneficiariosFound: list[Beneficiario] = self.sacConnector.getPossibleBeneficiarios(rutBeneficiario=rutBeneficiario)
        if beneficiariosFound:
            self.nombreBeneficiarioDropdown.config(values=[beneficiario.nombreBeneficiario for beneficiario in beneficiariosFound])
        else:
            self.nombreBeneficiarioDropdown.set('Sin resultados')

    def selectCaso(self, key=None):
        dataSelected: list = self.casosTable.item(self.casosTable.focus())['values']
        if dataSelected:
            rutDeudor: str = dataSelected[4]
            apellidoDeudor: str = dataSelected[5]
            nombreDeudor: str = dataSelected[6]
            nombreCliente: str = dataSelected[7]
            indexCliente: int = ([cliente.nombreCliente for cliente in self.clientes] + ['Ninguno']).index(nombreCliente)
            self.clienteDropdown.current(newindex=indexCliente)
            self.rutDeudorEntry.delete(0, END)
            self.rutDeudorEntry.insert(0, rutDeudor)
            self.nombreDeudorEntry.delete(0, END)
            self.nombreDeudorEntry.insert(0, nombreDeudor)
            self.apellidoDeudorEntry.delete(0, END)
            self.apellidoDeudorEntry.insert(0, apellidoDeudor)
            rendicionNumero = self.getClienteRendicion()
            self.rendicionEntry.delete(0, END)
            self.rendicionEntry.insert(0, rendicionNumero)
            self.setDestinatario()
    
    def beginClearForm(self):
        if not messagebox.askyesno(title='Aviso', message='Se borrarán todos los datos ingresados. \n ¿Desea continuar?'):
            return
        self.clearForm()
        
    def clearCacheFiles(self):
        deleteFileIfExists('thumbnail.png')
        deleteFileIfExists('result.png')
        
    def onClosingWindow(self):
        self.clearCacheFiles()
        self.master.destroy()
        self.master.update()
    
    def clearForm(self):
        self.boletaPath = ''
        self.anexosPaths.clear()
        self.numBoletaEntry.delete(0, END)
        self.fechaBoletaEntry.delete(0, END)
        self.rutBeneficiarioEntry.delete(0, END)
        self.nombreBeneficiarioDropdown.set('')
        self.rutDeudorEntry.delete(0, END)
        self.apellidoDeudorEntry.delete(0, END)
        self.nombreDeudorEntry.delete(0, END)
        self.clienteDropdown.set('')
        self.gastoTotalEntry.delete(0, END)
        self.casosTable.delete(*self.casosTable.get_children())
        self.serviciosTable.delete(*self.serviciosTable.get_children())
        self.uploadedAnexosLabel.config(text='No se han subido anexos')
        self.uploadedBoletaLabel.config(text='No se ha subido boleta')
        self.clearCacheFiles()

        self.master.destroy()
        self.master.update()
        app: App = App()