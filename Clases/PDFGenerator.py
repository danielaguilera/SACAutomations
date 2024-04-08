from fpdf import FPDF
from Clases.BoletaMatrixRow import BoletaMatrixRow
from Clases.Resumen import Resumen
from Clases.SACConnector import SACConnector
from Utils.GlobalFunctions import *
from Clases.Servicio import Servicio
from Utils.Metadata import *
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Border, Side, Alignment, NamedStyle

class PDFGenerator:
    def __init__(self):
        self.sacConnector = SACConnector()
        
    def generateReporte(self, resumenBoleta: Resumen):
        contactoReporte : str = resumenBoleta.destinatario.nombreDestinatario
        if contactoReporte.find('.'):
            prefijo: str = contactoReporte[0: contactoReporte.find('.') + 1]
            encabezado: str = contactoReporte[contactoReporte.find('.') + 2::]
        else:
            prefijo: str = ''
            encabezado: str = contactoReporte
        
        pdf: FPDF = FPDF('P', 'mm', 'Letter')

        pdf.add_page()
        pdf.set_font('helvetica', 'BI', 10)
        pdf.set_text_color(0, 0, 255)
        pdf.set_xy(x=130, y=10)
        pdf.cell(40, 10, transformDateToSpanish(date=datetime.now()))
        pdf.image(LOGOPATH, x=18, y=15, w=60, h=20)

        pdf.set_xy(x=18, y=40)
        pdf.set_font('Times', 'BI', 11)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(40, 10, prefijo)
        pdf.set_xy(x=18, y=48)
        pdf.cell(40, 10, encabezado)
        pdf.set_xy(x=18, y=56)
        pdf.cell(40, 10, resumenBoleta.cliente.factura)
        pdf.set_xy(x=18, y=64)
        pdf.cell(40, 10, 'Presente')
        pdf.set_xy(x=18, y=75)
        pdf.cell(40, 10, 'Sírvase tener por acompañadas, las siguientes boletas por los servicios que en cada caso se señala, así')
        pdf.set_xy(x=18, y=80)
        pdf.cell(40, 10, 'como los antecedentes que las justifican en cada caso.')
        
        pdf.set_xy(x=18, y=90)
        pdf.set_font(size=13, family='Arial')
        pdf.cell(40, 10, 'Beneficiario')
        pdf.set_font(size=11, family='Arial')
        pdf.set_xy(x=50, y=90)
        pdf.cell(40, 10, resumenBoleta.beneficiario.nombreBeneficiario)
        pdf.set_xy(x=140, y=90)
        pdf.set_font(size=13, family='Arial')
        pdf.cell(40, 10, 'RUT')
        pdf.set_font(size=14, family='Arial')
        pdf.set_xy(x=155, y=90)
        pdf.cell(40, 10, resumenBoleta.beneficiario.rutBeneficiario)

        pdf.set_font('helvetica', 'BI', 10)
        pdf.set_text_color(0, 0, 255)

        pdf.set_xy(x=18, y=100)
        pdf.cell(40, 10, 'RUT Deudor')
        pdf.set_xy(x=45, y=100)
        pdf.cell(40, 10, 'Apellido Deudor')
        pdf.set_xy(x=80, y=100)
        pdf.cell(40, 10, 'Nombre Deudor')
        pdf.set_xy(x=115, y=100)
        pdf.cell(40, 10, 'Boleta')
        pdf.set_xy(x=130, y=100)
        pdf.cell(40, 10, 'Fecha')
        pdf.set_xy(x=150, y=100)
        pdf.cell(40, 10, 'Monto')
        pdf.set_xy(x=170, y=100)
        pdf.cell(40, 10, 'Nota')

        pdf.line(x1=18, y1=108, x2=200, y2=108)

        servicio : Servicio
        delta = 13
        pdf.set_font(size=11, family='Arial')
        pdf.set_text_color(0, 0, 0)
        for index, servicio in enumerate(resumenBoleta.servicios):
            pdf.set_xy(x=18, y=106 + index*delta)
            pdf.cell(40, 10, resumenBoleta.caso.rutDeudor)
            pdf.set_font(size=9, family='Arial')
            pdf.set_xy(x=45, y=106 + index*delta)
            pdf.cell(40, 10, resumenBoleta.caso.apellidoDeudor)
            pdf.set_xy(x=80, y=106 + index*delta)
            pdf.cell(40, 10, resumenBoleta.caso.nombreDeudor)
            pdf.set_font(size=11, family='Arial')
            pdf.set_xy(x=115, y=106 + index*delta)
            pdf.cell(40, 10, str(resumenBoleta.boleta.numBoleta))
            pdf.set_xy(x=130, y=106 + index*delta)
            pdf.cell(40, 10, transformDateToSpanishBrief(date=resumenBoleta.boleta.fechaEmision))
            pdf.set_xy(x=150, y=106 + index*delta)
            pdf.cell(40, 10, setPriceFormat(servicio.monto))
            pdf.set_xy(x=170, y=106 + index*delta)
            pdf.set_font(size=8, family='Arial')
            pdf.cell(40, 10, servicio.nota)
            pdf.set_text_color(0, 0, 255)
            pdf.set_font('helvetica', 'BI', 10)
            pdf.set_xy(x=18, y=112 + index*delta)
            pdf.cell(40, 10, 'Cliente: ')
            pdf.set_xy(x=105, y=112 + index*delta)
            pdf.cell(40, 10, 'Código: ')
            pdf.set_font(size=11, family='Arial')
            pdf.set_text_color(0, 0, 0)
            pdf.set_xy(x=45, y=112 + index*delta)
            pdf.cell(40, 10, f'{str(resumenBoleta.cliente.idCliente)}    | {resumenBoleta.cliente.nombreCliente}')
            pdf.set_xy(x=120, y=112 + index*delta)
            pdf.cell(40, 10, servicio.codigo)
            pdf.line(x1=18, y1=120 + index*delta, x2=200, y2=120 + index*delta)

        pdf.set_xy(x=18, y=120 + index*delta)
        pdf.set_font(size=13, family='Arial')
        pdf.set_text_color(0, 0, 20)
        pdf.cell(40, 10, 'Suma total')
        pdf.set_xy(x=170, y=120 + index*delta)
        pdf.cell(40, 10, setPriceFormat(sum(servicio.monto for servicio in resumenBoleta.servicios)))
        pdf.image(SIGNINGPATH, x=18, y=150 + index*delta, w=140, h=40)        

        if not os.path.exists(f'{DELIVEREDDATAPATH}/{resumenBoleta.destinatario.nombreDestinatario}/{resumenBoleta.boleta.numBoleta}_{resumenBoleta.caso.idMapsa}'):
            os.makedirs(f'{DELIVEREDDATAPATH}/{resumenBoleta.destinatario.nombreDestinatario}/{resumenBoleta.boleta.numBoleta}_{resumenBoleta.caso.idMapsa}')
        pdf.output(f'{DELIVEREDDATAPATH}/{resumenBoleta.destinatario.nombreDestinatario}/{resumenBoleta.boleta.numBoleta}_{resumenBoleta.caso.idMapsa}/Reporte_{resumenBoleta.boleta.numBoleta}.pdf')
        self.updateExcelMatrix(nombreDestinatario=resumenBoleta.destinatario.nombreDestinatario,
                               nombreCliente=resumenBoleta.cliente.nombreCliente)
        
    def updateExcelMatrix(self, nombreDestinatario: str, nombreCliente: str):
        numeroRendicion = self.sacConnector.getRendicionNumber(nombreDestinatario=nombreDestinatario, nombreCliente=nombreCliente)
        excelMatrixRoot: str = f'{DELIVEREDDATAPATH}/{nombreDestinatario}/Rendición_{nombreCliente}_{numeroRendicion}.xlsx'
        deleteFileIfExists(excelMatrixRoot)
        self.createExcelMatrix(nombreDestinatario=nombreDestinatario, nombreCliente=nombreCliente)
        conn: SACConnector = SACConnector()
        for dirName in os.listdir(path=f'{DELIVEREDDATAPATH}/{nombreDestinatario}'):
            if dirName[0] in ['R', 'D']: # AQUI SE DEBEN PONER SOLO DEL CLIENTE
                 continue
            numBoleta, idMapsa = dirName.split('_')
            numBoleta = int(numBoleta)
            idMapsa = int(idMapsa)
            boletaNombreCliente = conn.getClienteFromCasoId(idMapsa=idMapsa).nombreCliente
            if boletaNombreCliente == nombreCliente:
                self.addBoletaDataToExcelMatrix(nombreDestinatario=nombreDestinatario, nombreCliente=nombreCliente, numBoleta=numBoleta)
        self.updateMatrixTotalValues(nombreDestinatario=nombreDestinatario, nombreCliente=nombreCliente, numeroRendicion=numeroRendicion)
        
    def createExcelMatrix(self, nombreDestinatario: str, nombreCliente: str):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        numeroRendicion = self.sacConnector.getRendicionNumber(nombreDestinatario=nombreDestinatario, nombreCliente=nombreCliente)
        headers = [' NOMBRE DEUDOR ', ' OPERACIÓN ', ' FOLIO ', ' ITEM ', ' SERVICIO ', ' RUT PRESTADOR ', ' PRESTADOR ', ' N°DOC ', ' MONTO ', ' N°BOLETA ', ' FECHA PAGO ']
        cell = worksheet.cell(row = 1, column = 1)
        pattern = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type='solid')
        cell.fill = pattern
        cell.value = nombreCliente.upper()
        cell = worksheet.cell(row = 1, column = 2)
        cell.fill = pattern
        cell.value = 'Reembolso - MVSERVICIOS'
        worksheet.merge_cells(f'B1:G1')
        cell = worksheet.cell(row = 1, column = 8)
        cell.fill = pattern
        cell.value = f'R N° {numeroRendicion}/0-0'
        worksheet.merge_cells(f'H1:I1')
        cell = worksheet.cell(row = 1, column = 10)
        cell.fill = pattern
        cell.value = datetime.strftime(datetime.now(), '%d-%m-%Y')
        worksheet.merge_cells(f'J1:K1')
        worksheet.merge_cells(f'A2:K2')
        i: bool = False
        
        for colNum, header in enumerate(headers):
            cell = worksheet.cell(row=3, column=colNum + 1)
            cell.value = header
            patternColor = SKY_BLUE if i else RED
            fontColor = BLACK if i else WHITE
            pattern = PatternFill(start_color=patternColor, end_color=patternColor, fill_type='solid')
            font = Font(color=fontColor)
            cell.font = font
            cell.fill = pattern
            i = not i
        workbook.save(f'{DELIVEREDDATAPATH}/{nombreDestinatario}/Rendición_{nombreCliente}_{numeroRendicion}.xlsx')
        
    def addBoletaDataToExcelMatrix(self, nombreDestinatario: str, nombreCliente: str, numBoleta: int):
        numeroRendicion = self.sacConnector.getRendicionNumber(nombreDestinatario=nombreDestinatario, nombreCliente=nombreCliente)
        excelMatrixRoot: str = f'{DELIVEREDDATAPATH}/{nombreDestinatario}/Rendición_{nombreCliente}_{numeroRendicion}.xlsx'
        if not os.path.exists(excelMatrixRoot):
            self.createExcelMatrix(nombreDestinatario=nombreDestinatario, nombreCliente=nombreCliente)
        workbook = openpyxl.load_workbook(excelMatrixRoot)
        sheet: openpyxl.worksheet.worksheet.Worksheet = workbook.active
        conn: SACConnector = SACConnector()
        rows: list[BoletaMatrixRow] = conn.getBoletaMatrixRows(numBoleta=numBoleta)
        filaNDoc = max([x for x in [sheet.cell(row=row, column=8).value for row in range(4, sheet.max_row + 1)] if x] + [0]) + 1
        for row in rows:
            filaNombreDeudor = row.nombreDeudor.upper()
            filaOperacion = row.nOperacion
            filaFolio = row.nFolio
            filaItem = row.item
            filaServicio = row.nombreServicio
            filaRUTPrestador = row.rutPrestador
            filaPrestador = row.nombrePrestador
            filaMonto = row.monto
            filaNBoleta = row.nBoleta
            filaFechaPago = row.fechaPago
            newRow = [filaNombreDeudor, 
                    filaOperacion, 
                    filaFolio, 
                    filaItem, 
                    filaServicio, 
                    filaRUTPrestador, 
                    filaPrestador, 
                    filaNDoc, 
                    filaMonto, 
                    filaNBoleta, 
                    filaFechaPago]
            sheet.append(newRow)
        
        border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'))
        
        currencyStyle = 'currencyStyle'
        if currencyStyle not in workbook.named_styles:
            style = NamedStyle(name=currencyStyle, number_format=" _-$* #,##0_-")
            workbook.add_named_style(style)
        
        for columnIndex, column in enumerate(sheet.columns):
            header = str(column[0].value)
            maxLength = len(header)
            column = [cell for cell in column]
            for cell in column:
                if len(str(cell.value)) > maxLength:
                    maxLength = len(cell.value)
                cell.border = border
            adjustedWidth = maxLength + 10
            sheet.column_dimensions[openpyxl.utils.get_column_letter(cell.column)].width = adjustedWidth
            for rowIndex, cell in enumerate(column):
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                if rowIndex > 2 and columnIndex == 3:
                    color = MATRIX_COLORS.get(str(cell.value).split('-')[0], '000000')
                    pattern = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    cell.fill = pattern
                if rowIndex > 2 and columnIndex in [0,4,6]:
                    alignment = Alignment(horizontal='left')
                    cell.alignment = alignment
                if rowIndex > 2 and columnIndex == 8:
                    cell.style = currencyStyle
                    cell.border = border
        nServicios: int = len(rows)
        maxRow: int = sheet.max_row
        sheet.merge_cells(f'H{maxRow - nServicios + 1}:H{maxRow}')
        sheet.merge_cells(f'J{maxRow - nServicios + 1}:J{maxRow}')
        sheet.merge_cells(f'K{maxRow - nServicios + 1}:K{maxRow}')
        workbook.save(excelMatrixRoot)

    def updateMatrixTotalValues(self, nombreDestinatario: str, nombreCliente: str, numeroRendicion: int = 0):
        excelMatrixRoot: str = f'{DELIVEREDDATAPATH}/{nombreDestinatario}/Rendición_{nombreCliente}_{numeroRendicion}.xlsx'
        if not os.path.exists(excelMatrixRoot):
            self.createExcelMatrix(nombreDestinatario=nombreDestinatario, nombreCliente=nombreCliente)
        workbook = openpyxl.load_workbook(excelMatrixRoot)
        sheet: openpyxl.worksheet.worksheet.Worksheet = workbook.active
        nBoletas = max([x for x in [sheet.cell(row=row, column=8).value for row in range(4, sheet.max_row + 1)] if x] + [0])
        cell = sheet.cell(row = 1, column = 8)
        cell.value = f'R N° {numeroRendicion}/1-{nBoletas}'
        workbook.save(excelMatrixRoot)
        workbook = openpyxl.load_workbook(excelMatrixRoot)
        sheet: openpyxl.worksheet.worksheet.Worksheet = workbook.active
        total = sum([sheet.cell(row = i, column = 9).value for i in range(4, sheet.max_row + 1)])
        newRow = ['TOTAL', '', '', '', '', '', '', '', total, '', '']
        sheet.append(newRow)
        workbook.save(excelMatrixRoot)
        workbook = openpyxl.load_workbook(excelMatrixRoot)
        sheet: openpyxl.worksheet.worksheet.Worksheet = workbook.active
        currencyStyle = 'currencyStyle'
        if currencyStyle not in workbook.named_styles:
            style = NamedStyle(name=currencyStyle, number_format=" _-$* #,##0_-")
            workbook.add_named_style(style)
        pattern = PatternFill(start_color=BLUE, end_color=BLUE, fill_type='solid')
        font = Font(color=WHITE)
        for j in range(1, 12):
            cell = sheet.cell(row = sheet.max_row, column = j)
            if j == 9:
                cell.style = currencyStyle
            cell.fill = pattern
            cell.font = font
        workbook.save(excelMatrixRoot)    
    
        
            
            
        

            